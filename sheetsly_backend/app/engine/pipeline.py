"""Ingestion and inspection pipeline orchestrator."""

from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Optional, Tuple

from app.core.errors import DatasetNotFoundError, SheetNotFoundError
from app.core.logging import logger
from app.engine.parser.sheet_reader import RawSheetGrid
from app.engine.parser.workbook_inspector import WorkbookInspector
from app.engine.profiler.quality_engine import DataQualityEngine
from app.engine.profiler.table_detector import TableDetector
from app.models.schemas import (
    DataQualityReport,
    SheetDataGridResponse,
    SheetMetadata,
    WorkbookOverview,
)


class IngestionPipeline:
    """Coordinates parsing, profiling, and metadata caching for workbook datasets."""

    def __init__(self):
        # In-memory store for active session datasets
        self._overview_cache: Dict[str, WorkbookOverview] = {}
        self._grids_cache: Dict[str, Dict[str, RawSheetGrid]] = {}

    def process_workbook(
        self,
        dataset_id: str,
        file_path: Path,
        original_filename: str,
        file_size_bytes: int,
    ) -> WorkbookOverview:
        """
        Executes complete deterministic inspection on a workbook file.
        Produces full structural metadata, table boundaries, orientation, and quality reports.
        """
        logger.info(f"Starting inspection pipeline for dataset {dataset_id} ({original_filename})")

        sheet_names, grids = WorkbookInspector.inspect_file(file_path)
        sheets_meta = []
        quality_scores = []

        for idx, name in enumerate(sheet_names):
            grid = grids[name]

            # 1. Detect candidate tables
            tables = TableDetector.detect_tables_in_sheet(grid)

            # 2. Evaluate data quality
            quality_report = DataQualityEngine.evaluate_sheet_quality(grid, tables)
            quality_scores.append(quality_report.overall_score)

            # 3. Compute empty rows & cols
            empty_rows = sum(1 for r in range(grid.min_row, grid.max_row + 1) if grid.is_row_empty(r))
            empty_cols = sum(1 for c in range(grid.min_col, grid.max_col + 1) if grid.is_col_empty(c))

            # 4. Count formula cells
            formula_count = sum(1 for cell in grid.cells.values() if cell.formula is not None)

            dimensions = f"{grid.used_range}"

            sheet_meta = SheetMetadata(
                name=name,
                index=idx,
                is_hidden=grid.is_hidden,
                dimensions=dimensions,
                total_rows=grid.total_rows,
                total_columns=grid.total_cols,
                used_range=grid.used_range,
                empty_rows_count=empty_rows,
                empty_cols_count=empty_cols,
                merged_cells_regions=grid.merged_ranges,
                formula_cells_count=formula_count,
                tables=tables,
                quality_report=quality_report,
            )
            sheets_meta.append(sheet_meta)

        overall_score = (
            round(sum(quality_scores) / len(quality_scores), 1) if quality_scores else 100.0
        )

        overview = WorkbookOverview(
            dataset_id=dataset_id,
            filename=original_filename,
            file_size_bytes=file_size_bytes,
            sheet_count=len(sheets_meta),
            sheets=sheets_meta,
            overall_quality_score=overall_score,
            created_at=datetime.now(timezone.utc).isoformat(),
        )

        # Cache session data
        self._overview_cache[dataset_id] = overview
        self._grids_cache[dataset_id] = grids

        logger.info(
            f"Dataset {dataset_id} processed successfully: {len(sheets_meta)} sheets, overall quality {overall_score}/100"
        )
        return overview

    def get_overview(self, dataset_id: str) -> WorkbookOverview:
        """Retrieves cached overview for a dataset ID."""
        if dataset_id not in self._overview_cache:
            raise DatasetNotFoundError(dataset_id)
        return self._overview_cache[dataset_id]

    def get_sheet_grid(self, dataset_id: str, sheet_name: str) -> RawSheetGrid:
        """Retrieves raw grid for a specific sheet in a dataset."""
        if dataset_id not in self._grids_cache:
            raise DatasetNotFoundError(dataset_id)
        grids = self._grids_cache[dataset_id]
        if sheet_name not in grids:
            raise SheetNotFoundError(sheet_name)
        return grids[sheet_name]

    def get_sheet_data_page(
        self,
        dataset_id: str,
        sheet_name: str,
        page: int = 1,
        page_size: int = 50,
        search_query: Optional[str] = None,
    ) -> SheetDataGridResponse:
        """
        Retrieves a paginated 2D cell slice of actual spreadsheet data for the frontend viewer.
        Preserves row numbers, column headers, cell coordinates, and raw/parsed values.
        Supports case-insensitive keyword search across all cells with accurate matching row pagination.
        """
        grid = self.get_sheet_grid(dataset_id, sheet_name)

        total_rows = grid.total_rows
        total_cols = grid.total_cols

        if total_rows == 0 or total_cols == 0:
            return SheetDataGridResponse(
                dataset_id=dataset_id,
                sheet_name=sheet_name,
                page=page,
                page_size=page_size,
                total_rows=0,
                total_columns=0,
                column_headers=[],
                rows=[],
                merged_cells=grid.merged_ranges,
            )

        from openpyxl.utils import get_column_letter

        column_headers = [get_column_letter(c) for c in range(grid.min_col, grid.max_col + 1)]

        clean_q = search_query.strip().lower() if search_query and search_query.strip() else None

        if clean_q:
            # Deterministic multi-column row search
            matching_row_indices = []
            for r in range(grid.min_row, grid.max_row + 1):
                row_matches = False
                for c in range(grid.min_col, grid.max_col + 1):
                    cell = grid.get_cell(r, c)
                    if cell.original_value is not None and clean_q in str(cell.original_value).lower():
                        row_matches = True
                        break
                    if cell.parsed_value is not None and clean_q in str(cell.parsed_value).lower():
                        row_matches = True
                        break
                    if cell.formula and clean_q in cell.formula.lower():
                        row_matches = True
                        break
                if row_matches:
                    matching_row_indices.append(r)

            total_matching_rows = len(matching_row_indices)
            start_offset = (page - 1) * page_size
            end_offset = min(total_matching_rows, start_offset + page_size)
            selected_rows = matching_row_indices[start_offset:end_offset]

            rows_slice = []
            for r in selected_rows:
                row_cells = [grid.get_cell(r, c) for c in range(grid.min_col, grid.max_col + 1)]
                rows_slice.append(row_cells)

            return SheetDataGridResponse(
                dataset_id=dataset_id,
                sheet_name=sheet_name,
                page=page,
                page_size=page_size,
                total_rows=total_matching_rows,
                total_columns=total_cols,
                column_headers=column_headers,
                rows=rows_slice,
                merged_cells=grid.merged_ranges,
            )

        # Standard unsearched pagination
        start_row_idx = grid.min_row + (page - 1) * page_size
        end_row_idx = min(grid.max_row, start_row_idx + page_size - 1)

        rows_slice = []
        if start_row_idx <= grid.max_row:
            for r in range(start_row_idx, end_row_idx + 1):
                row_cells = [grid.get_cell(r, c) for c in range(grid.min_col, grid.max_col + 1)]
                rows_slice.append(row_cells)

        return SheetDataGridResponse(
            dataset_id=dataset_id,
            sheet_name=sheet_name,
            page=page,
            page_size=page_size,
            total_rows=total_rows,
            total_columns=total_cols,
            column_headers=column_headers,
            rows=rows_slice,
            merged_cells=grid.merged_ranges,
        )


ingestion_pipeline = IngestionPipeline()

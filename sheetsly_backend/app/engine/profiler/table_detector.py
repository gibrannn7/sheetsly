"""Table detector for extracting structured table regions, columns, and lineage bounds."""

import uuid
from typing import List, Tuple
from openpyxl.utils import get_column_letter

from app.engine.parser.sheet_reader import RawSheetGrid
from app.models.schemas import (
    ColumnMetadata,
    OrientationEnum,
    TableRegion,
)
from .orientation_detector import OrientationDetector
from .region_detector import ClassifiedRegion, RegionDetector
from .type_detector import TypeDetector


class TableDetector:
    """Detects and profiles candidate table regions within a worksheet."""

    @classmethod
    def detect_tables_in_sheet(cls, grid: RawSheetGrid) -> List[TableRegion]:
        """
        Extracts all candidate TableRegion objects from a RawSheetGrid.
        """
        if grid.total_rows == 0 or grid.total_cols == 0:
            return []

        regions = RegionDetector.detect_regions(grid)
        tables: List[TableRegion] = []

        for idx, reg in enumerate(regions, start=1):
            min_r, min_c, max_r, max_c = reg.bounding_box

            # Evaluate orientation
            orient_result = OrientationDetector.detect_orientation(grid, min_r, min_c, max_r, max_c)

            # Build table bounds
            start_col_letter = get_column_letter(min_c)
            end_col_letter = get_column_letter(max_c)
            table_range = f"{start_col_letter}{min_r}:{end_col_letter}{max_r}"

            header_range = None
            data_range = None

            if reg.header_rows:
                h_min = min(reg.header_rows)
                h_max = max(reg.header_rows)
                header_range = f"{start_col_letter}{h_min}:{end_col_letter}{h_max}"

            if reg.data_rows:
                d_min = min(reg.data_rows)
                d_max = max(reg.data_rows)
                data_range = f"{start_col_letter}{d_min}:{end_col_letter}{d_max}"
            elif reg.header_rows and max_r > max(reg.header_rows):
                # Fallback data range
                d_min = max(reg.header_rows) + 1
                data_range = f"{start_col_letter}{d_min}:{end_col_letter}{max_r}"

            # Extract Columns Profile
            columns_meta: List[ColumnMetadata] = []
            seen_names: dict[str, int] = {}

            for c_pos, col_idx in enumerate(reg.active_cols):
                col_letter = get_column_letter(col_idx)

                # Extract header name
                header_cell_ref = None
                header_parts = []
                for h_row in reg.header_rows:
                    h_cell = grid.get_cell(h_row, col_idx)
                    val = h_cell.original_value
                    if val is not None and str(val).strip() != "":
                        header_parts.append(str(val).strip())
                        if not header_cell_ref:
                            header_cell_ref = h_cell.coordinate.cell_ref

                raw_name = " / ".join(header_parts) if header_parts else f"Column {col_letter}"

                # Disambiguate duplicate column names
                if raw_name in seen_names:
                    seen_names[raw_name] += 1
                    col_name = f"{raw_name}_{seen_names[raw_name]}"
                else:
                    seen_names[raw_name] = 1
                    col_name = raw_name

                # Gather data values for profiling
                data_values = []
                data_rows_to_use = reg.data_rows or (
                    list(range(max(reg.header_rows) + 1, max_r + 1)) if reg.header_rows else []
                )
                for d_row in data_rows_to_use:
                    d_cell = grid.get_cell(d_row, col_idx)
                    data_values.append(d_cell.original_value)

                # Profile column
                dt, sem, conf, nulls, uniq, samples = TypeDetector.profile_column_vector(
                    data_values, column_name=col_name
                )

                columns_meta.append(
                    ColumnMetadata(
                        index=c_pos,
                        name=col_name,
                        original_header_cell=header_cell_ref,
                        source_column_letter=col_letter,
                        data_type=dt,
                        semantic_type=sem,
                        type_confidence=conf,
                        total_count=len(data_values),
                        null_count=nulls,
                        unique_count=uniq,
                        sample_values=samples,
                    )
                )

            data_row_count = len(reg.data_rows) if reg.data_rows else max(0, max_r - (max(reg.header_rows) if reg.header_rows else min_r))
            col_count = len(reg.active_cols)

            table_id = f"tbl_{grid.sheet_name.lower().replace(' ', '_')}_{idx}"
            table_name = f"{grid.sheet_name} Table {idx}" if len(regions) > 1 else f"{grid.sheet_name} Data"

            # Compute overall table confidence
            # High if header exists, data rows exist, and orientation is clear
            table_conf = 0.95
            if not reg.header_rows:
                table_conf -= 0.2
            if data_row_count == 0:
                table_conf -= 0.3
            if orient_result.orientation == OrientationEnum.AMBIGUOUS:
                table_conf -= 0.15

            tables.append(
                TableRegion(
                    table_id=table_id,
                    name=table_name,
                    sheet_name=grid.sheet_name,
                    range_address=table_range,
                    header_range=header_range,
                    data_range=data_range,
                    header_row_indices=reg.header_rows,
                    orientation=orient_result.orientation,
                    orientation_confidence=orient_result.confidence,
                    orientation_reasons=orient_result.reasons,
                    row_count=data_row_count,
                    column_count=col_count,
                    columns=columns_meta,
                    confidence_score=round(max(0.1, min(1.0, table_conf)), 3),
                )
            )

        return tables

"""Authoritative Rollback Engine for atomic restoration of spreadsheet grid state."""

import copy
from typing import Any, Dict, List, Optional, Tuple
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter

from app.engine.agent.action_model import ActionTypeEnum, SpreadsheetAction
from app.engine.agent.transaction_model import (
    CellSnapshot,
    MutationTransaction,
    StateDiff,
)
from app.engine.parser.sheet_reader import RawSheetGrid
from app.models.schemas import CellCoordinate, CellData, DataTypeEnum


class RollbackEngine:
    """Executes atomic restoration of spreadsheet grid state from snapshots and diffs."""

    @classmethod
    def rollback_transaction(
        cls,
        transaction: MutationTransaction,
        grid: RawSheetGrid,
        sheet_grids: Optional[Dict[str, RawSheetGrid]] = None,
        deep_backup_grids: Optional[Dict[str, RawSheetGrid]] = None,
    ) -> bool:
        """
        Reverses all actions executed in a transaction, ensuring CURRENT_STATE == BEFORE_STATE.
        """
        try:
            # 1. If deep backup grids were provided, restore directly from full in-memory deepcopy
            if deep_backup_grids:
                if grid.sheet_name in deep_backup_grids:
                    backup = deep_backup_grids[grid.sheet_name]
                    grid.cells = copy.deepcopy(backup.cells)
                    grid.min_row = backup.min_row
                    grid.max_row = backup.max_row
                    grid.min_col = backup.min_col
                    grid.max_col = backup.max_col
                    grid.charts = copy.deepcopy(getattr(backup, "charts", {}))
                    grid.kpis = copy.deepcopy(getattr(backup, "kpis", {}))
                if sheet_grids:
                    for s_name, s_grid in sheet_grids.items():
                        if s_name in deep_backup_grids:
                            backup = deep_backup_grids[s_name]
                            s_grid.cells = copy.deepcopy(backup.cells)
                            s_grid.min_row = backup.min_row
                            s_grid.max_row = backup.max_row
                            s_grid.min_col = backup.min_col
                            s_grid.max_col = backup.max_col
                            s_grid.charts = copy.deepcopy(getattr(backup, "charts", {}))
                            s_grid.kpis = copy.deepcopy(getattr(backup, "kpis", {}))
                return True

            # 2. Reverse action-by-action rollback
            for act in reversed(transaction.actions):
                target_grid = grid
                if sheet_grids and act.sheet_name in sheet_grids:
                    target_grid = sheet_grids[act.sheet_name]

                if act.action_type in {ActionTypeEnum.WRITE_VALUE, ActionTypeEnum.WRITE_FORMULA, ActionTypeEnum.FORMAT_CELL, ActionTypeEnum.SET_NUMBER_FORMAT}:
                    if act.target_cell:
                        cls._restore_cell_from_snapshot(target_grid, act.target_cell, transaction.before_state)

                elif act.action_type == ActionTypeEnum.FORMAT_RANGE:
                    if act.target_range:
                        cls._restore_range_from_snapshot(target_grid, act.target_range, transaction.before_state)

                elif act.action_type == ActionTypeEnum.CLEAR_CONTENT:
                    if act.target_cell:
                        cls._restore_cell_from_snapshot(target_grid, act.target_cell, transaction.before_state)
                    elif act.target_range:
                        cls._restore_range_from_snapshot(target_grid, act.target_range, transaction.before_state)

                elif act.action_type == ActionTypeEnum.INSERT_ROW and act.row_index:
                    cls._reverse_insert_row(target_grid, act.row_index)

                elif act.action_type == ActionTypeEnum.INSERT_COLUMN and act.column_index:
                    cls._reverse_insert_column(target_grid, act.column_index)

                elif act.action_type in {ActionTypeEnum.CREATE_CHART, ActionTypeEnum.UPDATE_CHART} and act.chart_spec:
                    target_grid.charts.pop(act.chart_spec.chart_id, None)

                elif act.action_type == ActionTypeEnum.MOVE_CHART and act.chart_spec:
                    if act.chart_spec.chart_id in target_grid.charts:
                        # Restore previous destination
                        target_grid.charts[act.chart_spec.chart_id]["destination_cell"] = act.chart_spec.destination_cell

                elif act.action_type == ActionTypeEnum.DELETE_CHART and act.chart_spec:
                    target_grid.charts[act.chart_spec.chart_id] = act.chart_spec.model_dump()

                elif act.action_type == ActionTypeEnum.CREATE_KPI and act.kpi_spec:
                    target_grid.kpis.pop(act.kpi_spec.kpi_id, None)

                elif act.action_type == ActionTypeEnum.CREATE_WORKSHEET:
                    if sheet_grids and act.sheet_name in sheet_grids:
                        sheet_grids.pop(act.sheet_name, None)

            return True
        except Exception:
            return False

    @classmethod
    def _restore_cell_from_snapshot(
        cls,
        grid: RawSheetGrid,
        cell_ref: str,
        before_state: Dict[str, CellSnapshot],
    ):
        col_str, row_int = coordinate_from_string(cell_ref.upper())
        col_int = column_index_from_string(col_str)

        if cell_ref.upper() in before_state:
            snap = before_state[cell_ref.upper()]
            dt = DataTypeEnum(snap.data_type) if snap.data_type in DataTypeEnum._value2member_map_ else DataTypeEnum.UNKNOWN
            grid.cells[(row_int, col_int)] = CellData(
                coordinate=CellCoordinate(row=row_int, column=col_int, cell_ref=cell_ref.upper(), col_letter=col_str),
                original_value=snap.original_value,
                parsed_value=snap.parsed_value,
                data_type=dt,
                formula=snap.formula,
                is_empty=snap.is_empty,
            )
        else:
            # Was not in before_state, meaning it was newly created / empty
            if (row_int, col_int) in grid.cells:
                grid.cells.pop((row_int, col_int), None)

    @classmethod
    def _restore_range_from_snapshot(
        cls,
        grid: RawSheetGrid,
        range_ref: str,
        before_state: Dict[str, CellSnapshot],
    ):
        if ":" in range_ref:
            s_ref, e_ref = range_ref.split(":")
            s_col_str, s_row = coordinate_from_string(s_ref)
            e_col_str, e_row = coordinate_from_string(e_ref)
            s_col = column_index_from_string(s_col_str)
            e_col = column_index_from_string(e_col_str)

            for r in range(min(s_row, e_row), max(s_row, e_row) + 1):
                for c in range(min(s_col, e_col), max(s_col, e_col) + 1):
                    c_let = get_column_letter(c)
                    c_ref = f"{c_let}{r}"
                    cls._restore_cell_from_snapshot(grid, c_ref, before_state)

    @classmethod
    def _reverse_insert_row(cls, grid: RawSheetGrid, row_idx: int):
        new_cells: Dict[Tuple[int, int], CellData] = {}
        for (r, c), cell in grid.cells.items():
            if r > row_idx:
                new_r = r - 1
                col_let = get_column_letter(c)
                new_coord = CellCoordinate(row=new_r, column=c, cell_ref=f"{col_let}{new_r}", col_letter=col_let)
                new_cells[(new_r, c)] = cell.model_copy(update={"coordinate": new_coord})
            elif r < row_idx:
                new_cells[(r, c)] = cell
            # row_idx itself is dropped as it was the inserted row
        grid.cells = new_cells
        grid.max_row = max(1, grid.max_row - 1)

    @classmethod
    def _reverse_insert_column(cls, grid: RawSheetGrid, col_idx: int):
        new_cells: Dict[Tuple[int, int], CellData] = {}
        for (r, c), cell in grid.cells.items():
            if c > col_idx:
                new_c = c - 1
                col_let = get_column_letter(new_c)
                new_coord = CellCoordinate(row=r, column=new_c, cell_ref=f"{col_let}{r}", col_letter=col_let)
                new_cells[(r, new_c)] = cell.model_copy(update={"coordinate": new_coord})
            elif c < col_idx:
                new_cells[(r, c)] = cell
            # col_idx itself is dropped as it was the inserted column
        grid.cells = new_cells
        grid.max_col = max(1, grid.max_col - 1)

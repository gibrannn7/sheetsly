import re
import uuid
from typing import Any, Dict, List, Optional, Tuple
import pandas as pd
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter

from app.engine.agent.action_model import (
    ActionTypeEnum,
    ChartActionSpec,
    FormattingStyle,
    KPIActionSpec,
    SpreadsheetAction,
)
from app.engine.agent.action_validator import ActionValidator
from app.engine.agent.formula_evaluator import FormulaEvaluator
from app.engine.agent.placement_policy import PlacementPolicy
from app.engine.agent.transaction_model import AgentResponseStatusEnum
from app.engine.ai.models import ClarificationRequest
from app.engine.analytics.aggregations import DeterministicAggregator
from app.engine.analytics.ambiguity_resolver import GeneralizedAmbiguityResolver
from app.engine.analytics.instruction_model import OperationEnum
from app.engine.parser.sheet_reader import RawSheetGrid
from app.engine.profiler.workbook_index import (
    ColumnIndexEntry,
    SheetIndexEntry,
    TableIndexEntry,
    WorkbookMetadataIndex,
)
from app.engine.visualization.chart_model import ChartSeriesSpec, ChartTypeEnum
from app.engine.visualization.renderer import ChartRenderer
from app.models.schemas import DataTypeEnum, SemanticTypeEnum
from app.storage.file_manager import file_manager


class SpreadsheetAgentPlanner:
    """Plans canonical spreadsheet actions from natural language user instructions."""

    @classmethod
    def plan_agent_actions(
        cls,
        user_request: str,
        workbook_index: WorkbookMetadataIndex,
        grid: RawSheetGrid,
        active_sheet_name: Optional[str] = None,
        selected_range: Optional[str] = None,
    ) -> Tuple[List[SpreadsheetAction], AgentResponseStatusEnum, Optional[ClarificationRequest], str]:
        """
        Translates user instructions into an atomic, validated action sequence with explicit
        cell/range targeting, safe formatting, and deterministic Python verification.
        """
        cur_sheet = active_sheet_name or workbook_index.active_sheet_name
        sheet_entry = workbook_index.sheets.get(cur_sheet)
        if not sheet_entry or not sheet_entry.tables:
            return [], AgentResponseStatusEnum.VALIDATION_ERROR, None, f"Sheet '{cur_sheet}' does not contain structured tables."

        table_entry = sheet_entry.tables[0]
        q_norm = user_request.strip().lower()
        q_clean = re.sub(r'[^\w\s]', ' ', q_norm).strip()
        q_tokens = [w for w in q_clean.split() if w]

        is_english_query = any(w in q_tokens for w in ["what", "this", "dataset", "data", "explain", "describe", "contain", "calculate", "sum", "average", "in", "of"]) and not any(w in q_tokens for w in ["ini", "apa", "tentang", "jelaskan", "jelasin", "sebenarnya", "isinya", "hitung", "buatkan", "tulis", "selesai", "di", "ke", "pada", "rentang"])

        # Defense-in-depth: If transaction control command reaches planner, immediately return without mutations
        if any(w in q_tokens for w in ["undo", "redo", "cancel", "rollback", "revert"]) or any(kw in q_clean for kw in [
            "undo langkah", "batalkan langkah", "batalkan perubahan", "kembalikan seperti", "ulangi langkah", "ulangi perubahan",
            "terapkan kembali", "batalkan operasi", "jangan lakukan", "batalkan itu", "undo perubahan", "undo aksi"
        ]):
            return [], AgentResponseStatusEnum.SUCCESS, None, "Transaction control operation detected."

        # Check for explicit placement/destination in prompt:
        # Extract explicit destination cell (e.g. "di N2", "ke N20", "D10", "in N20", "at D10")
        dest_match = re.search(
            r'(?:di\s+rentang|pada\s+rentang|rentang|range|di\s+range|pada\s+range|mulai\s+dari|mulai|start\s+from|start\s+at|dari\s+sel|di\s+posisi|posisi|position|di\s+lokasi|lokasi|location|letakkan\s+(?:di|pada)?|tempatkan\s+(?:di|pada)?|place\s+at|put\s+at|taruh\s+(?:di|pada)?|didalam|di\s+dalam|di|in|at|pada|target|ke|to|sel|cell|di\s+sel|pada\s+sel|in\s+cell|at\s+cell|area|di\s+area|pada\s+area|kotak|di\s+kotak|pada\s+kotak)\s+([A-Za-z]{1,3}\d+)(?::([A-Za-z]{1,3}\d+))?\b',
            q_norm
        )
        explicit_dest_cell = dest_match.group(1).upper() if dest_match else None
        dest_range_str = f"{dest_match.group(1).upper()}:{dest_match.group(2).upper()}" if (dest_match and dest_match.group(2)) else None

        # Check for standalone coordinate in request (e.g. "Tebalkan D10", "Format D10 sebagai currency")
        if not explicit_dest_cell:
            cell_m = re.search(r'\b([A-Za-z]{1,3}\d+)\b', user_request)
            if cell_m:
                explicit_dest_cell = cell_m.group(1).upper()

        # Contextual reference to selected cell / range (e.g. "di cell ini", "di cells ini", "di sini", "at this cell", "here")
        has_contextual_dest_kw = any(kw in q_norm for kw in [
            "di cell ini", "di cells ini", "pada cell ini", "pada cells ini", "sel ini", "di sel ini",
            "di sini", "disini", "di posisi ini", "di tempat ini", "posisi ini", "pada posisi ini",
            "at this cell", "in this cell", "here", "at the selected cell", "in selected cell",
            "di cell yang dipilih", "pada cell terpilih", "cell yang dipilih", "sel terpilih"
        ])
        if has_contextual_dest_kw and selected_range:
            selected_single = selected_range.split(":")[0].upper()
            if not explicit_dest_cell:
                explicit_dest_cell = selected_single

        # If selected_range is a single cell (e.g. "B40") and no explicit coordinate was given:
        is_single_selected = bool(selected_range and ":" not in selected_range and re.match(r'^[A-Za-z]{1,3}\d+$', selected_range.strip()))
        if not explicit_dest_cell and is_single_selected:
            has_source_kw = any(kw in q_norm for kw in ["dari data yang dipilih", "dari range yang dipilih", "dari data terpilih", "from selected data", "from selection"])
            if not has_source_kw:
                explicit_dest_cell = selected_range.strip().upper()

        # Check for explicit source data range in prompt or selected_range
        source_range_match = re.search(r'(?:dari\s+rentang|dari\s+range|dari\s+data|dari\s+|from\s+range|from\s+data|from\s+|source|sumber)\s*([A-Za-z]{1,3}\d+:[A-Za-z]{1,3}\d+)\b', q_norm)
        explicit_ranges = re.findall(r'\b([A-Za-z]{1,3}\d+:[A-Za-z]{1,3}\d+)\b', user_request)

        if source_range_match:
            source_range_context = source_range_match.group(1).upper()
        elif selected_range and ":" in selected_range:
            source_range_context = selected_range.upper()
        elif explicit_ranges:
            # Filter out any range that was matched as destination placement
            non_dest = [r.upper() for r in explicit_ranges if r.upper() != dest_range_str]
            source_range_context = non_dest[0] if non_dest else None
        else:
            source_range_context = None

        if source_range_context and explicit_dest_cell:
            if source_range_context.startswith(explicit_dest_cell) or explicit_dest_cell in source_range_context.split(":"):
                # The matched cell was actually part of the source range (e.g. "dari C2:C11")
                explicit_dest_cell = None

        target_range_context = source_range_context

        # Extract explicit multi-intent label specification
        label_match_1 = re.search(
            r'(?:dengan\s+label|berikan\s+label|tulis\s+label|label(?:\s+it)?|put\s+label|with\s+label|put)\s+([\'\"].*?[\'\"]|[a-zA-Z0-9_\s\-]+?)\s+(?:di|pada|ke|in|at|on|to)\s+([A-Za-z]{1,3}\d+)\b',
            user_request,
            re.IGNORECASE
        )
        label_match_2 = re.search(
            r'(?:di\s+|pada\s+|ke\s+|in\s+|at\s+)?([A-Za-z]{1,3}\d+)\s+(?:berikan\s+label|tulis\s+label|dengan\s+label|isi\s+label|put\s+label|label)\s+([\'\"].*?[\'\"]|[a-zA-Z0-9_\s\-]+)',
            user_request,
            re.IGNORECASE
        )

        explicit_label_text = None
        explicit_label_cell = None

        if label_match_1:
            cand_txt = label_match_1.group(1).strip().strip("'\"")
            if cand_txt.lower() not in ["currency", "persen", "bold", "formula", "sales", "profit", "discount", "quantity"]:
                explicit_label_text = cand_txt
                explicit_label_cell = label_match_1.group(2).upper()
        elif label_match_2:
            cand_txt = label_match_2.group(2).strip().strip("'\"")
            if cand_txt.lower() not in ["currency", "persen", "bold", "formula", "sales", "profit", "discount", "quantity"]:
                explicit_label_cell = label_match_2.group(1).upper()
                explicit_label_text = cand_txt

        # Disambiguate explicit_dest_cell and explicit_label_cell
        all_coords = [c.upper() for c in re.findall(r'\b[A-Za-z]{1,3}\d+\b', user_request)]
        if explicit_label_cell and explicit_dest_cell == explicit_label_cell:
            other_coords = [c for c in all_coords if c != explicit_label_cell]
            explicit_dest_cell = other_coords[0] if other_coords else None

        # Check for explicit formatting request inside multi-intent query
        explicit_format_type = None
        explicit_format_cell = None
        format_match = re.search(
            r'(?:and\s+|dan\s+)?(?:format\s+([A-Za-z]{1,3}\d+)?\s*(?:sebagai|as)\s*(currency|mata\s+uang|rupiah|dollar|persen|percentage|bold|tebal)|tebalkan\s+([A-Za-z]{1,3}\d+)?)',
            q_norm
        )
        if format_match:
            fmt_target = format_match.group(1) or format_match.group(3)
            fmt_kind = format_match.group(2)
            explicit_format_cell = fmt_target.upper() if fmt_target else None
            if fmt_kind in ["currency", "mata uang", "rupiah", "dollar"]:
                explicit_format_type = "currency"
            elif fmt_kind in ["persen", "percentage"]:
                explicit_format_type = "percentage"
            else:
                explicit_format_type = "bold"

        # -------------------------------------------------------------
        # 1. READ-ONLY DATASET EXPLANATION (0 Mutations)
        # -------------------------------------------------------------
        dataset_explain_patterns = [
            r'ini\s+(?:sebenarnya\s+)?data\s+(?:apa|tentang\s+apa|apasih)',
            r'data\s+(?:apa\s+)?(?:ini\s+)?(?:tentang\s+apa|isinya\s+apa|apa\s*sih|apa|apasih)',
            r'dataset\s+(?:ini\s+)?(?:tentang\s+apa|isinya\s+apa|apa|apasih)',
            r'(?:jelas(?:kan|in)|deskripsi(?:kan)?|terang(?:kan)?|rangkum)\s+(?:isi\s+)?(?:data|dataset)',
            r'apa\s+(?:sebenarnya\s+)?(?:isi\s+)?data(?:set)?',
            r'what\s+(?:is\s+this\s+data|is\s+this\s+dataset|does\s+this\s+data\s+contain)',
            r'explain\s+(?:this\s+)?data(?:set)?',
            r'describe\s+(?:this\s+)?data(?:set)?',
            r'summarize\s+(?:this\s+)?data(?:set)?',
        ]
        is_explain_dataset = any(re.search(pat, q_clean) for pat in dataset_explain_patterns) or (
            any(w in q_tokens for w in ["data", "dataset"]) and
            any(w in q_tokens for w in ["apa", "apasih", "tentang", "jelaskan", "jelasin", "explain", "describe", "isinya", "isi", "sebenarnya"]) and
            not any(w in q_tokens for w in ["total", "sum", "hitung", "rata-rata", "average", "min", "max", "chart", "grafik", "tulis", "format", "hapus", "clear"])
        )

        if is_explain_dataset:
            total_rows = getattr(table_entry, 'row_count', None) or getattr(table_entry, 'total_rows', None) or getattr(sheet_entry, 'total_rows', 0)
            total_cols = len(table_entry.columns)

            id_cols = [c.name for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.IDENTIFIER]
            cat_cols = [c.name for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.CATEGORICAL or c.data_type == DataTypeEnum.STRING]
            temp_cols = [c.name for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.TEMPORAL or c.data_type in {DataTypeEnum.DATE, DataTypeEnum.DATETIME}]
            num_cols = [c.name for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.INTEGER, DataTypeEnum.PERCENTAGE}]
            missing_cols = [f"{c.name} ({getattr(c, 'missing_percentage', 0):.1f}%)" for c in table_entry.columns if (getattr(c, 'missing_percentage', 0) or 0) > 0]

            table_desc_name = table_entry.name.replace("_Table", "").replace("_", " ") if table_entry.name else cur_sheet

            if is_english_query:
                parts = [
                    f"This is a {table_desc_name} dataset with {total_rows:,} rows and {total_cols} columns."
                ]
                details = []
                if id_cols:
                    details.append(f"identifiers like {', '.join(id_cols[:3])}")
                if cat_cols:
                    details.append(f"categories such as {', '.join(cat_cols[:4])}")
                if temp_cols:
                    details.append(f"dates like {', '.join(temp_cols)}")
                if num_cols:
                    details.append(f"metrics such as {', '.join(num_cols[:4])}")
                if details:
                    parts.append(f"The data contains {', '.join(details)}.")
                if missing_cols:
                    parts.append(f"Columns with missing values: {', '.join(missing_cols[:3])}.")
                else:
                    parts.append("Data quality is clean with no significant missing values.")
            else:
                parts = [
                    f"Ini adalah dataset {table_desc_name} dengan {total_rows:,} baris dan {total_cols} kolom."
                ]
                details = []
                if id_cols:
                    details.append(f"informasi {', '.join(id_cols[:3])}")
                if cat_cols:
                    details.append(f"kategori seperti {', '.join(cat_cols[:4])}")
                if temp_cols:
                    details.append(f"tanggal transaksi ({', '.join(temp_cols)})")
                if num_cols:
                    details.append(f"metrik seperti {', '.join(num_cols[:4])}")
                if details:
                    parts.append(f"Data mencakup {', '.join(details)}.")
                if missing_cols:
                    parts.append(f"Beberapa kolom memiliki nilai kosong: {', '.join(missing_cols[:3])}.")
                else:
                    parts.append("Kualitas data bersih tanpa nilai kosong signifikan.")

            return [], AgentResponseStatusEnum.SUCCESS, None, " ".join(parts)

        # -------------------------------------------------------------
        # 2. READ-ONLY CELL INSPECTION (0 Mutations)
        # -------------------------------------------------------------
        cell_inspect_match = re.search(r'(?:apa\s+(?:isi|yang\s+ada\s+di)|jelas(?:kan|in)|inspect|what\s+is\s+in|lihat\s+isi)\s+(?:sel\s+|cell\s+)?([A-Za-z]{1,3}\d+)\b', q_clean)
        is_inspect_cell = bool(cell_inspect_match) or any(kw in q_clean for kw in [
            "apa isi cell", "apa isi sel", "isi cell ini", "isi sel ini", "jelaskan cell", "jelaskan sel", "nilai cell", "nilai sel", "what is in this cell", "inspect cell", "explain cell", "cek isi sel", "cek isi cell", "lihat isi sel", "apa yang ada di sel", "apa yang ada di cell"
        ])

        if is_inspect_cell:
            inspect_coord = (
                (cell_inspect_match.group(1).upper() if cell_inspect_match else None) or
                explicit_dest_cell or
                (selected_range.split(":")[0] if selected_range else None) or
                "A1"
            ).upper()
            try:
                col_s, row_i = coordinate_from_string(inspect_coord)
                col_i = column_index_from_string(col_s)
                c_data = grid.get_cell(row_i, col_i)
                d_type = c_data.data_type.value if hasattr(c_data.data_type, 'value') else str(c_data.data_type)
                if c_data.is_empty or (c_data.parsed_value is None and not c_data.formula):
                    msg = f"Sel {inspect_coord} saat ini kosong."
                elif c_data.formula:
                    msg = f"Sel {inspect_coord} berisi formula '{c_data.formula}' dengan nilai terurai '{c_data.parsed_value}' (tipe: {d_type})."
                else:
                    msg = f"Sel {inspect_coord} berisi nilai '{c_data.parsed_value}' (tipe: {d_type})."
                return [], AgentResponseStatusEnum.SUCCESS, None, msg
            except Exception as e:
                return [], AgentResponseStatusEnum.VALIDATION_ERROR, None, f"Gagal membaca sel {inspect_coord}: {str(e)}"

        # -------------------------------------------------------------
        # 4. READ-ONLY / DIRECT CONTEXT-AWARE ANALYTICAL INQUIRIES
        # -------------------------------------------------------------
        # A. Cross-Dimensional Partitioning (e.g. "category dengan sales tertinggi di setiap region")
        cross_dim_match = re.search(
            r'(?:category|kategori)\s+(?:dengan|yang)\s+(?:sales|penjualan|profit)\s+(?:tertinggi|terbesar|paling\s+tinggi|highest)\s+(?:di\s+setiap|pada\s+setiap|per|in\s+each)\s+(?:region|wilayah)',
            q_clean
        ) or re.search(
            r'(?:highest|top)\s+(?:sales\s+)?(?:category)\s+per\s+region',
            q_clean
        )

        if cross_dim_match:
            reg_col = next((c for c in table_entry.columns if c.normalized_name == "region" or "region" in c.normalized_name), None)
            cat_col = next((c for c in table_entry.columns if c.normalized_name == "category" or "category" in c.normalized_name), None)
            sales_col = next((c for c in table_entry.columns if c.normalized_name in ["sales", "profit"] or c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE), None)

            if reg_col and cat_col and sales_col:
                start_r = 2
                end_r = getattr(table_entry, 'row_count', 5) + 1
                if table_entry.data_range and ":" in table_entry.data_range:
                    try:
                        start_r = int(re.search(r'\d+', table_entry.data_range.split(":")[0]).group())
                        end_r = int(re.search(r'\d+', table_entry.data_range.split(":")[1]).group())
                    except Exception:
                        pass

                reg_c_idx = column_index_from_string(reg_col.source_column_letter)
                cat_c_idx = column_index_from_string(cat_col.source_column_letter)
                sales_c_idx = column_index_from_string(sales_col.source_column_letter)

                reg_cat_sales: Dict[str, Dict[str, float]] = {}
                for r in range(start_r, end_r + 1):
                    reg_val = str(grid.get_cell(r, reg_c_idx).parsed_value or "Unknown")
                    cat_val = str(grid.get_cell(r, cat_c_idx).parsed_value or "Unknown")
                    s_val = grid.get_cell(r, sales_c_idx).parsed_value
                    try:
                        s_float = float(s_val) if s_val is not None else 0.0
                    except Exception:
                        s_float = 0.0
                    if reg_val not in reg_cat_sales:
                        reg_cat_sales[reg_val] = {}
                    reg_cat_sales[reg_val][cat_val] = reg_cat_sales[reg_val].get(cat_val, 0.0) + s_float

                top_per_reg = []
                for reg, cats in reg_cat_sales.items():
                    if cats:
                        best_cat = max(cats.items(), key=lambda x: x[1])
                        top_per_reg.append(f"{reg}: {best_cat[0]} (${best_cat[1]:,.2f})")

                if is_english_query:
                    msg = f"Category with the highest {sales_col.name} per {reg_col.name}: {', '.join(top_per_reg)}."
                else:
                    msg = f"Kategori dengan {sales_col.name} tertinggi di setiap {reg_col.name}: {', '.join(top_per_reg)}."

                if explicit_dest_cell:
                    action = SpreadsheetAction(
                        action_id="act_1",
                        action_type=ActionTypeEnum.WRITE_VALUE,
                        sheet_name=cur_sheet,
                        target_cell=explicit_dest_cell,
                        value="; ".join(top_per_reg),
                        description=f"Write highest category per region to {explicit_dest_cell}",
                    )
                    return [action], AgentResponseStatusEnum.SUCCESS, None, msg

                return [], AgentResponseStatusEnum.SUCCESS, None, msg

        # B. Superlative Dimension Ranking (e.g. "region mana dengan sales tertinggi?", "region mana dengan sales terendah?")
        dim_rank_match = re.search(
            r'(region|category|wilayah|kategori)\s+(?:mana\s+)?(?:dengan|yang)?\s+(?:sales|penjualan|profit|keuntungan)\s+(tertinggi|terbesar|paling\s+tinggi|highest|top|terendah|terkecil|paling\s+rendah|lowest|bottom)',
            q_clean
        ) or re.search(
            r'(?:which|what)\s+(region|category)\s+has\s+(?:the\s+)?(highest|top|lowest|bottom)\s+(sales|profit)',
            q_clean
        )

        if dim_rank_match:
            is_min = any(w in q_clean for w in ["terendah", "terkecil", "paling rendah", "lowest", "bottom"])
            dim_name = dim_rank_match.group(1).lower()
            target_dim_col = next((c for c in table_entry.columns if dim_name in c.normalized_name or c.normalized_name in dim_name or c.semantic_type == SemanticTypeEnum.CATEGORICAL), None)
            sales_col = next((c for c in table_entry.columns if c.normalized_name in ["sales", "profit"] or c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE), None)

            if target_dim_col and sales_col:
                start_r = 2
                end_r = getattr(table_entry, 'row_count', 5) + 1
                if table_entry.data_range and ":" in table_entry.data_range:
                    try:
                        start_r = int(re.search(r'\d+', table_entry.data_range.split(":")[0]).group())
                        end_r = int(re.search(r'\d+', table_entry.data_range.split(":")[1]).group())
                    except Exception:
                        pass

                dim_c_idx = column_index_from_string(target_dim_col.source_column_letter)
                sales_c_idx = column_index_from_string(sales_col.source_column_letter)

                dim_sums: Dict[str, float] = {}
                for r in range(start_r, end_r + 1):
                    d_val = str(grid.get_cell(r, dim_c_idx).parsed_value or "Unknown")
                    s_val = grid.get_cell(r, sales_c_idx).parsed_value
                    try:
                        s_float = float(s_val) if s_val is not None else 0.0
                    except Exception:
                        s_float = 0.0
                    dim_sums[d_val] = dim_sums.get(d_val, 0.0) + s_float

                if dim_sums:
                    if is_min:
                        best_item = min(dim_sums.items(), key=lambda x: x[1])
                        rank_lbl_id = "terendah"
                        rank_lbl_en = "lowest"
                    else:
                        best_item = max(dim_sums.items(), key=lambda x: x[1])
                        rank_lbl_id = "tertinggi"
                        rank_lbl_en = "highest"

                    if is_english_query:
                        msg = f"The {target_dim_col.name} with the {rank_lbl_en} total {sales_col.name} is {best_item[0]} (${best_item[1]:,.2f})."
                    else:
                        msg = f"{target_dim_col.name} dengan total {sales_col.name} {rank_lbl_id} adalah {best_item[0]} (${best_item[1]:,.2f})."

                    if explicit_dest_cell:
                        action = SpreadsheetAction(
                            action_id="act_1",
                            action_type=ActionTypeEnum.WRITE_VALUE,
                            sheet_name=cur_sheet,
                            target_cell=explicit_dest_cell,
                            value=best_item[0],
                            description=f"Write {rank_lbl_en} {target_dim_col.name} to {explicit_dest_cell}",
                        )
                        return [action], AgentResponseStatusEnum.SUCCESS, None, msg

                    return [], AgentResponseStatusEnum.SUCCESS, None, msg

        # C. Scalar Extremes (e.g. "berapa sales terbesar?", "what is the highest sales?")
        scalar_extreme_match = re.search(
            r'(?:berapa\s+|what\s+is\s+the\s+)?(sales|profit|quantity|diskon|discount)\s+(?:terbesar|tertinggi|highest|maximum|max|terkecil|terendah|lowest|minimum|min)',
            q_clean
        )
        if scalar_extreme_match and not any(w in q_clean for w in ["tulis", "write", "hitung", "buatkan", "tampilkan"]):
            is_min = any(w in q_clean for w in ["terkecil", "terendah", "lowest", "minimum", "min"])
            metric_term = scalar_extreme_match.group(1).lower()
            m_col = next((c for c in table_entry.columns if metric_term in c.normalized_name or c.normalized_name in metric_term), None)
            if m_col:
                start_r = 2
                end_r = getattr(table_entry, 'row_count', 5) + 1
                if table_entry.data_range and ":" in table_entry.data_range:
                    try:
                        start_r = int(re.search(r'\d+', table_entry.data_range.split(":")[0]).group())
                        end_r = int(re.search(r'\d+', table_entry.data_range.split(":")[1]).group())
                    except Exception:
                        pass
                m_c_idx = column_index_from_string(m_col.source_column_letter)
                vals = []
                for r in range(start_r, end_r + 1):
                    v = grid.get_cell(r, m_c_idx).parsed_value
                    try:
                        if v is not None:
                            vals.append(float(v))
                    except Exception:
                        pass
                if vals:
                    ext_val = min(vals) if is_min else max(vals)
                    op_lbl_id = "terkecil" if is_min else "terbesar"
                    op_lbl_en = "lowest" if is_min else "highest"
                    if is_english_query:
                        msg = f"The {op_lbl_en} {m_col.name} value is ${ext_val:,.2f}."
                    else:
                        msg = f"{m_col.name} {op_lbl_id} adalah ${ext_val:,.2f}."

                    if explicit_dest_cell:
                        action = SpreadsheetAction(
                            action_id="act_1",
                            action_type=ActionTypeEnum.WRITE_VALUE,
                            sheet_name=cur_sheet,
                            target_cell=explicit_dest_cell,
                            value=ext_val,
                            description=f"Write {op_lbl_en} {m_col.name} to {explicit_dest_cell}",
                        )
                        return [action], AgentResponseStatusEnum.SUCCESS, None, msg

                    return [], AgentResponseStatusEnum.SUCCESS, None, msg

        # -------------------------------------------------------------
        # 5. CONTROLLED WORKBOOK COMPLETION ("selesaikan data ini", "complete this analysis")
        # -------------------------------------------------------------
        completion_patterns = [
            r'selesaikan\s+(?:analisis\s+)?data(?:\s+ini)?',
            r'lengkapi\s+(?:analisis\s+)?data(?:\s+ini)?',
            r'complete\s+(?:this\s+)?(?:analysis|data)',
            r'prepare\s+(?:this\s+)?data\s+for\s+analysis',
        ]
        is_completion_req = any(re.search(pat, q_clean) for pat in completion_patterns)

        if is_completion_req:
            start_r = 2
            end_r = getattr(table_entry, 'row_count', 5) + 1
            if table_entry.data_range and ":" in table_entry.data_range:
                try:
                    start_r = int(re.search(r'\d+', table_entry.data_range.split(":")[0]).group())
                    end_r = int(re.search(r'\d+', table_entry.data_range.split(":")[1]).group())
                except Exception:
                    pass

            summary_start_r = end_r + 2
            measures = [c for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.INTEGER}]

            actions: List[SpreadsheetAction] = []
            act_id = 1

            # Header Row
            hdr_text = "Analysis Summary" if is_english_query else "Ringkasan Analisis"
            actions.append(
                SpreadsheetAction(
                    action_id=f"act_{act_id}",
                    action_type=ActionTypeEnum.WRITE_VALUE,
                    sheet_name=cur_sheet,
                    target_cell=f"A{summary_start_r}",
                    value=hdr_text,
                    description=f"Write summary header to A{summary_start_r}",
                )
            )
            act_id += 1
            actions.append(
                SpreadsheetAction(
                    action_id=f"act_{act_id}",
                    action_type=ActionTypeEnum.FORMAT_RANGE,
                    sheet_name=cur_sheet,
                    target_range=f"A{summary_start_r}:B{summary_start_r}",
                    style=FormattingStyle(bold=True, fill_color="#F1F5F9"),
                    description=f"Format summary header row at A{summary_start_r}:B{summary_start_r}",
                )
            )
            act_id += 1

            curr_r = summary_start_r + 1
            if measures:
                primary_m = measures[0]
                # Total primary measure
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{act_id}",
                        action_type=ActionTypeEnum.WRITE_VALUE,
                        sheet_name=cur_sheet,
                        target_cell=f"A{curr_r}",
                        value=f"Total {primary_m.name}",
                        description=f"Write label 'Total {primary_m.name}' to A{curr_r}",
                    )
                )
                act_id += 1
                formula_total = f"=SUM({primary_m.source_column_letter.upper()}{start_r}:{primary_m.source_column_letter.upper()}{end_r})"
                exp_total, _ = FormulaEvaluator.evaluate(formula_total, grid)
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{act_id}",
                        action_type=ActionTypeEnum.WRITE_FORMULA,
                        sheet_name=cur_sheet,
                        target_cell=f"B{curr_r}",
                        formula=formula_total,
                        expected_result=exp_total,
                        description=f"Write formula '{formula_total}' to B{curr_r}",
                    )
                )
                act_id += 1
                curr_r += 1

                # Average primary measure
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{act_id}",
                        action_type=ActionTypeEnum.WRITE_VALUE,
                        sheet_name=cur_sheet,
                        target_cell=f"A{curr_r}",
                        value=f"Average {primary_m.name}" if is_english_query else f"Rata-rata {primary_m.name}",
                        description=f"Write label 'Average {primary_m.name}' to A{curr_r}",
                    )
                )
                act_id += 1
                formula_avg = f"=AVERAGE({primary_m.source_column_letter.upper()}{start_r}:{primary_m.source_column_letter.upper()}{end_r})"
                exp_avg, _ = FormulaEvaluator.evaluate(formula_avg, grid)
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{act_id}",
                        action_type=ActionTypeEnum.WRITE_FORMULA,
                        sheet_name=cur_sheet,
                        target_cell=f"B{curr_r}",
                        formula=formula_avg,
                        expected_result=exp_avg,
                        description=f"Write formula '{formula_avg}' to B{curr_r}",
                    )
                )
                act_id += 1
                curr_r += 1

                # If second measure exists (e.g. Profit)
                if len(measures) > 1:
                    second_m = measures[1]
                    actions.append(
                        SpreadsheetAction(
                            action_id=f"act_{act_id}",
                            action_type=ActionTypeEnum.WRITE_VALUE,
                            sheet_name=cur_sheet,
                            target_cell=f"A{curr_r}",
                            value=f"Total {second_m.name}",
                            description=f"Write label 'Total {second_m.name}' to A{curr_r}",
                        )
                    )
                    act_id += 1
                    formula_m2 = f"=SUM({second_m.source_column_letter.upper()}{start_r}:{second_m.source_column_letter.upper()}{end_r})"
                    exp_m2, _ = FormulaEvaluator.evaluate(formula_m2, grid)
                    actions.append(
                        SpreadsheetAction(
                            action_id=f"act_{act_id}",
                            action_type=ActionTypeEnum.WRITE_FORMULA,
                            sheet_name=cur_sheet,
                            target_cell=f"B{curr_r}",
                            formula=formula_m2,
                            expected_result=exp_m2,
                            description=f"Write formula '{formula_m2}' to B{curr_r}",
                        )
                    )
                    act_id += 1
                    curr_r += 1

            val_res = ActionValidator.validate_sequence(actions, workbook_index, grid)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."

            end_summary_r = curr_r - 1
            if is_english_query:
                msg = f"Done. Completed workbook analysis with summary metrics in A{summary_start_r}:B{end_summary_r}."
            else:
                msg = f"Selesai. Analisis data telah dilengkapi dengan ringkasan metrik di A{summary_start_r}:B{end_summary_r}."

            return actions, AgentResponseStatusEnum.SUCCESS, None, msg

        # -------------------------------------------------------------
        # 3. READ-ONLY COLUMN INQUIRY (0 Mutations)
        # -------------------------------------------------------------
        is_column_inquiry = any(kw in q_norm for kw in [
            "apa kolom yang bisa", "kolom apa yang bisa", "kolom mana yang cocok", "kolom apa yang cocok", "kolom untuk total", "kolom untuk tanggal", "which column"
        ])
        if is_column_inquiry:
            if any(kw in q_norm for kw in ["tanggal", "waktu", "date", "time", "tren"]):
                temp_cols = [c.name for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.TEMPORAL or c.data_type in {DataTypeEnum.DATE, DataTypeEnum.DATETIME}]
                if temp_cols:
                    msg = f"Kolom yang cocok untuk tanggal/waktu: {', '.join(temp_cols)}."
                else:
                    msg = "Tidak ditemukan kolom bertipe tanggal/waktu pada dataset ini."
            else:
                num_cols = [c.name for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.INTEGER, DataTypeEnum.PERCENTAGE}]
                if num_cols:
                    msg = f"Kolom numerik yang dapat digunakan untuk perhitungan total/rata-rata: {', '.join(num_cols)}."
                else:
                    msg = "Tidak ditemukan kolom metrik numerik pada dataset ini."
            return [], AgentResponseStatusEnum.SUCCESS, None, msg

        # Scope guard: Verify request relates to spreadsheet operations or table columns
        spreadsheet_keywords = [
            "total", "sum", "jumlah", "rata-rata", "average", "mean", "hitung", "count",
            "min", "max", "median", "buatkan", "buat", "tambahkan", "tambah", "tampilkan",
            "rekap", "summary", "format", "bold", "tebal", "tebalkan", "clear", "hapus",
            "sisipkan", "insert", "rumus", "formula", "baris", "kolom", "row", "column",
            "cell", "sel", "currency", "mata uang", "rupiah", "dollar", "persen", "percentage",
            "italic", "miring", "warna", "highlight", "header", "fill",
            "chart", "grafik", "diagram", "plot", "pie", "bar", "column", "line",
            "area", "scatter", "histogram", "visualisasi", "visualisasikan", "dashboard",
            "dasbor", "kpi", "data", "dataset", "isi", "calculate", "profit", "sales", "discount", "quantity", "cost"
        ]
        has_kw = any(kw in q_norm for kw in spreadsheet_keywords)
        has_col = any(c.normalized_name in q_norm or c.source_column_letter.lower() == q_norm for c in table_entry.columns)
        has_coord_in_prompt = bool(re.search(r'\b[A-Za-z]{1,3}\d+(?::[A-Za-z]{1,3}\d+)?\b', user_request))

        if not has_kw and not has_col and not has_coord_in_prompt and not selected_range and not user_request.strip().startswith("="):
            return [], AgentResponseStatusEnum.UNSUPPORTED, None, "Permintaan di luar cakupan operasi spreadsheet yang didukung."

        # -------------------------------------------------------------
        # BRANCH 0: Direct Write Formula / Write Value from Formula Bar or prompt
        # -------------------------------------------------------------
        direct_formula_match = re.search(
            r'(?:tulis\s+(?:rumus|formula)?\s*|write\s+formula\s+|isi\s+(?:rumus|formula)?\s*|^)(=.+?)(?:\s+(?:di|in|pada|ke|mulai|pada\s+sel|di\s+sel)\s+([A-Za-z]{1,3}\d+))?$',
            user_request.strip(),
            re.IGNORECASE
        )
        direct_write_pattern_1 = re.search(
            r'^(?:tulis|write|input|masukkan|isi)\s+(?:nilai\s+|teks\s+|angka\s+)?([\'"].*?[\'"]|\d+(?:\.\d+)?|[a-zA-Z0-9_\-]+)\s+(?:di|in|pada|ke|mulai|pada\s+sel|di\s+sel)\s+([A-Za-z]{1,3}\d+)',
            user_request.strip(),
            re.IGNORECASE
        )
        direct_write_pattern_2 = re.search(
            r'^(?:ganti\s+isi|ubah\s+isi|isi|set)\s+(?:sel\s+|cell\s+)?([A-Za-z]{1,3}\d+)\s+(?:dengan|menjadi|to|with|=)\s+([\'"].*?[\'"]|\d+(?:\.\d+)?|\S+)',
            user_request.strip(),
            re.IGNORECASE
        )

        if direct_formula_match and direct_formula_match.group(1).strip().startswith("="):
            formula_str = direct_formula_match.group(1).strip()
            target_c = (direct_formula_match.group(2) or explicit_dest_cell or (selected_range.split(":")[0] if selected_range else None) or "A1").upper()
            expected_val, _ = FormulaEvaluator.evaluate(formula_str, grid)
            action = SpreadsheetAction(
                action_id="act_1",
                action_type=ActionTypeEnum.WRITE_FORMULA,
                sheet_name=cur_sheet,
                target_cell=target_c,
                formula=formula_str,
                expected_result=expected_val,
                description=f"Write formula '{formula_str}' to cell {target_c}",
            )
            val_res = ActionValidator.validate_sequence([action], workbook_index, grid)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."
            return [action], AgentResponseStatusEnum.SUCCESS, None, f"Selesai. Rumus '{formula_str}' ditulis di {target_c}."

        if direct_write_pattern_1:
            val_raw = direct_write_pattern_1.group(1).strip().strip("'\"")
            try:
                val_parsed = float(val_raw) if "." in val_raw else int(val_raw)
            except ValueError:
                val_parsed = val_raw
            target_c = direct_write_pattern_1.group(2).upper()
            action = SpreadsheetAction(
                action_id="act_1",
                action_type=ActionTypeEnum.WRITE_VALUE,
                sheet_name=cur_sheet,
                target_cell=target_c,
                value=val_parsed,
                description=f"Write value '{val_parsed}' to cell {target_c}",
            )
            is_replace = any(kw in q_norm for kw in ["ganti", "replace", "overwrite", "timpa", "ubah"])
            val_res = ActionValidator.validate_sequence([action], workbook_index, grid, allow_overwrite=is_replace)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."
            return [action], AgentResponseStatusEnum.SUCCESS, None, f"Selesai. Nilai '{val_parsed}' ditulis di {target_c}."

        if direct_write_pattern_2:
            target_c = direct_write_pattern_2.group(1).upper()
            val_raw = direct_write_pattern_2.group(2).strip().strip("'\"")
            try:
                val_parsed = float(val_raw) if "." in val_raw else int(val_raw)
            except ValueError:
                val_parsed = val_raw
            action = SpreadsheetAction(
                action_id="act_1",
                action_type=ActionTypeEnum.WRITE_VALUE,
                sheet_name=cur_sheet,
                target_cell=target_c,
                value=val_parsed,
                description=f"Write value '{val_parsed}' to cell {target_c}",
            )
            val_res = ActionValidator.validate_sequence([action], workbook_index, grid, allow_overwrite=True)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."
            return [action], AgentResponseStatusEnum.SUCCESS, None, f"Selesai. Nilai '{val_parsed}' ditulis di {target_c}."

        # -------------------------------------------------------------
        # BRANCH V: Visualization, Dashboard & KPI Operations
        # -------------------------------------------------------------
        viz_keywords = [
            "chart", "grafik", "diagram", "plot", "pie", "bar chart", "column chart", "line chart",
            "area chart", "scatter", "histogram", "visualisasi", "visualisasikan", "dashboard",
            "dasbor", "kpi", "pie chart", "semua kemungkinan", "berikan semuanya", "tampilkan semuanya",
            "jangan hanya itu", "jangan cuma itu", "semua visualisasi", "semua chart", "semua grafik",
            "visualize", "visualizations", "all visualizations", "visualize everything", "visualize this data",
        ]
        is_viz_req = any(kw in q_norm for kw in viz_keywords)
        if is_viz_req:
            viz_res = cls._plan_visualization_actions(
                user_request=user_request,
                workbook_index=workbook_index,
                grid=grid,
                cur_sheet=cur_sheet,
                table_entry=table_entry,
                selected_range=selected_range,
                explicit_dest_cell=explicit_dest_cell,
                target_range_context=target_range_context,
            )
            if viz_res is not None:
                return viz_res

        # -------------------------------------------------------------
        # BRANCH A: Formatting Operations
        # -------------------------------------------------------------
        has_calc_intent = any(kw in q_norm for kw in [
            "total", "sum", "jumlah", "rata-rata", "average", "mean", "hitung", "count",
            "minimum", "maksimum", "selesaikan", "lengkapi"
        ])
        is_pure_formatting_req = any(kw in q_norm for kw in [
            "format", "bold", "tebal", "tebalkan", "currency", "mata uang",
            "rupiah", "dollar", "persen", "percentage", "italic", "miring", "warna", "highlight", "fill"
        ]) and not has_calc_intent
        is_header_req = any(kw in q_norm for kw in ["header", "judul", "baris atas", "top row"])

        if is_pure_formatting_req:
            actions: List[SpreadsheetAction] = []
            action_idx = 1

            if is_header_req:
                # Format Table Header Range
                header_range = table_entry.header_range
                if not header_range:
                    min_c = table_entry.columns[0].source_column_letter.upper() if table_entry.columns else "A"
                    max_c = table_entry.columns[-1].source_column_letter.upper() if table_entry.columns else "R"
                    header_range = f"{min_c}1:{max_c}1"

                target_range_to_format = header_range
                fmt_desc = "header tabel"
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{action_idx}",
                        action_type=ActionTypeEnum.FORMAT_RANGE,
                        sheet_name=cur_sheet,
                        target_range=header_range,
                        style=FormattingStyle(bold=True, fill_color="#F1F5F9"),
                        description=f"Format table header range '{header_range}' with bold styling and neutral fill",
                    )
                )
                action_idx += 1
            else:
                # Determine target column or range
                target_range_to_format = explicit_dest_cell or target_range_context
                if not target_range_to_format:
                    # Check if user mentioned a column
                    col_match = None
                    for col in table_entry.columns:
                        if col.normalized_name in q_norm or col.name.lower() in q_norm or col.source_column_letter.lower() == q_norm:
                            col_match = col
                            break
                    if col_match:
                        s_row = 2
                        e_row = 101
                        if table_entry.data_range and ":" in table_entry.data_range:
                            try:
                                _, s_row = coordinate_from_string(table_entry.data_range.split(":")[0])
                                _, e_row = coordinate_from_string(table_entry.data_range.split(":")[1])
                            except Exception:
                                pass
                        target_range_to_format = f"{col_match.source_column_letter.upper()}{s_row}:{col_match.source_column_letter.upper()}{e_row}"
                    elif selected_range:
                        target_range_to_format = selected_range.upper()

                if not target_range_to_format:
                    # Default to primary measure column range if available
                    measure_cols = [c for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE]
                    if measure_cols:
                        m_col = measure_cols[0]
                        s_row = 2
                        e_row = 101
                        if table_entry.data_range and ":" in table_entry.data_range:
                            try:
                                _, s_row = coordinate_from_string(table_entry.data_range.split(":")[0])
                                _, e_row = coordinate_from_string(table_entry.data_range.split(":")[1])
                            except Exception:
                                pass
                        target_range_to_format = f"{m_col.source_column_letter.upper()}{s_row}:{m_col.source_column_letter.upper()}{e_row}"
                    else:
                        target_range_to_format = "A1"

                # Check specific format style
                is_currency = any(kw in q_norm for kw in ["currency", "mata uang", "rupiah", "dollar", "$", "rp"])
                is_pct = any(kw in q_norm for kw in ["persen", "percentage", "%"])
                is_bold = any(kw in q_norm for kw in ["bold", "tebal", "tebalkan"])
                is_italic = any(kw in q_norm for kw in ["italic", "miring"])

                is_single_cell = ":" not in target_range_to_format

                if is_currency:
                    fmt_str = "Rp#,##0" if any(kw in q_norm for kw in ["rupiah", "rp", "idr"]) else "$#,##0.00"
                    actions.append(
                        SpreadsheetAction(
                            action_id=f"act_{action_idx}",
                            action_type=ActionTypeEnum.SET_NUMBER_FORMAT,
                            sheet_name=cur_sheet,
                            target_cell=target_range_to_format if is_single_cell else None,
                            target_range=None if is_single_cell else target_range_to_format,
                            number_format=fmt_str,
                            description=f"Apply currency number format '{fmt_str}' to {target_range_to_format}",
                        )
                    )
                    fmt_desc = f"currency ({fmt_str})"
                elif is_pct:
                    actions.append(
                        SpreadsheetAction(
                            action_id=f"act_{action_idx}",
                            action_type=ActionTypeEnum.SET_NUMBER_FORMAT,
                            sheet_name=cur_sheet,
                            target_cell=target_range_to_format if is_single_cell else None,
                            target_range=None if is_single_cell else target_range_to_format,
                            number_format="0.0%",
                            description=f"Apply percentage number format '0.0%' to {target_range_to_format}",
                        )
                    )
                    fmt_desc = "percentage"
                elif is_bold or is_italic:
                    actions.append(
                        SpreadsheetAction(
                            action_id=f"act_{action_idx}",
                            action_type=ActionTypeEnum.FORMAT_CELL if is_single_cell else ActionTypeEnum.FORMAT_RANGE,
                            sheet_name=cur_sheet,
                            target_cell=target_range_to_format if is_single_cell else None,
                            target_range=None if is_single_cell else target_range_to_format,
                            style=FormattingStyle(bold=is_bold, italic=is_italic),
                            description=f"Apply font style (bold={is_bold}, italic={is_italic}) to {target_range_to_format}",
                        )
                    )
                    fmt_desc = "tebal" if is_bold else "miring"
                else:
                    actions.append(
                        SpreadsheetAction(
                            action_id=f"act_{action_idx}",
                            action_type=ActionTypeEnum.FORMAT_CELL if is_single_cell else ActionTypeEnum.FORMAT_RANGE,
                            sheet_name=cur_sheet,
                            target_cell=target_range_to_format if is_single_cell else None,
                            target_range=None if is_single_cell else target_range_to_format,
                            style=FormattingStyle(bold=True),
                            description=f"Format range '{target_range_to_format}'",
                        )
                    )
                    fmt_desc = "format standar"

            val_res = ActionValidator.validate_sequence(actions, workbook_index, grid)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."
            return actions, AgentResponseStatusEnum.SUCCESS, None, f"Selesai. {target_range_to_format} diformat sebagai {fmt_desc}."

        # -------------------------------------------------------------
        # BRANCH B: Calculation / Formula Operations
        # -------------------------------------------------------------
        # 1. Determine aggregation function (SUM, AVERAGE, COUNT, MIN, MAX)
        agg_func = "SUM"
        agg_label = "Total"
        if any(kw in q_norm for kw in ["rata-rata", "average", "mean"]):
            agg_func = "AVERAGE"
            agg_label = "Average"
        elif any(kw in q_norm for kw in ["hitung baris", "count rows", "jumlah data", "jumlah baris", "jumlah record", "count", "banyaknya"]):
            agg_func = "COUNT"
            agg_label = "Count"
        elif any(kw in q_norm for kw in ["minimum", "min", "terkecil", "terendah"]):
            agg_func = "MIN"
            agg_label = "Min"
        elif any(kw in q_norm for kw in ["maksimum", "max", "terbesar", "tertinggi"]):
            agg_func = "MAX"
            agg_label = "Max"

        # 2. Strict Semantic Metric Resolution (Rule 1, 2, 3, 4)
        # Check if the user explicitly requested a specific metric
        SEMANTIC_METRIC_MAP = {
            "profit": "profit",
            "keuntungan": "profit",
            "laba": "profit",
            "untung": "profit",
            "sales": "sales",
            "penjualan": "sales",
            "omset": "sales",
            "omzet": "sales",
            "revenue": "sales",
            "pendapatan": "sales",
            "quantity": "quantity",
            "kuantitas": "quantity",
            "qty": "quantity",
            "discount": "discount",
            "diskon": "discount",
            "potongan": "discount",
            "cost": "cost",
            "biaya": "cost",
            "pengeluaran": "cost",
            "expense": "cost",
            "shipping cost": "shipping cost",
            "ongkir": "shipping cost",
        }

        # A. Check exact column name or normalized name in user request
        exact_matched_cols = [
            c for c in table_entry.columns
            if re.search(rf'\b{re.escape(c.normalized_name)}\b', q_norm) or re.search(rf'\b{re.escape(c.name.lower())}\b', q_norm)
        ]

        # B. Check semantic metric aliases
        alias_matched_cols = []
        requested_metric_keyword = None
        for alias_kw, standard_concept in SEMANTIC_METRIC_MAP.items():
            if re.search(rf'\b{re.escape(alias_kw)}\b', q_norm):
                requested_metric_keyword = requested_metric_keyword or alias_kw.title()
                for col in table_entry.columns:
                    if (
                        col.normalized_name == standard_concept or
                        standard_concept in col.normalized_name or
                        col.name.lower() == standard_concept or
                        standard_concept in col.name.lower()
                    ):
                        if col not in alias_matched_cols:
                            alias_matched_cols.append(col)

        target_col = None
        if exact_matched_cols:
            if len(exact_matched_cols) == 1:
                target_col = exact_matched_cols[0]
            else:
                req = ClarificationRequest(
                    question=f"Terdapat beberapa kolom yang cocok: {', '.join([c.name for c in exact_matched_cols])}. Kolom mana yang ingin digunakan?",
                    reason="Multiple candidate columns matched the query equally.",
                    target_parameter="target_column",
                    options=[c.name for c in exact_matched_cols],
                )
                return [], AgentResponseStatusEnum.CLARIFICATION, req, "Terdapat beberapa kolom yang cocok. Mohon pilih kolom yang diinginkan."
        elif alias_matched_cols:
            if len(alias_matched_cols) == 1:
                target_col = alias_matched_cols[0]
            else:
                req = ClarificationRequest(
                    question=f"Terdapat beberapa kolom yang cocok: {', '.join([c.name for c in alias_matched_cols])}. Kolom mana yang ingin digunakan?",
                    reason="Multiple candidate columns matched the query equally.",
                    target_parameter="target_column",
                    options=[c.name for c in alias_matched_cols],
                )
                return [], AgentResponseStatusEnum.CLARIFICATION, req, "Terdapat beberapa kolom yang cocok. Mohon pilih kolom yang diinginkan."
        elif requested_metric_keyword:
            # User explicitly requested a metric that DOES NOT EXIST in the workbook!
            # HARD RULE 4: NEVER FALL BACK SILENTLY TO SALES!
            available_num_cols = [c.name for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.INTEGER, DataTypeEnum.PERCENTAGE}]
            if is_english_query:
                avail_str = ", ".join(available_num_cols) if available_num_cols else "None"
                msg = f"I could not find the '{requested_metric_keyword}' column in this data. Available numeric columns are: {avail_str}."
            else:
                avail_str = ", ".join(available_num_cols) if available_num_cols else "tidak ada"
                msg = f"Saya tidak menemukan kolom '{requested_metric_keyword}' di data ini. Kolom numerik yang tersedia adalah {avail_str}."
            return [], AgentResponseStatusEnum.UNSUPPORTED, None, msg
        else:
            # User did NOT specify a metric (e.g. "hitung total" or "calculate total in D10")
            if target_range_context:
                start_c = target_range_context.split(":")[0]
                try:
                    col_letter, _ = coordinate_from_string(start_c)
                    for col in table_entry.columns:
                        if col.source_column_letter.upper() == col_letter.upper():
                            target_col = col
                            break
                except Exception:
                    pass

            if not target_col:
                measure_cols = [c for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.INTEGER}]
                if len(measure_cols) == 1:
                    target_col = measure_cols[0]
                elif len(measure_cols) > 1:
                    req = ClarificationRequest(
                        question=f"Ingin menghitung {agg_label} untuk kolom apa? ({', '.join([c.name for c in measure_cols])})",
                        reason="Multiple numeric measure columns available without explicit column in query.",
                        target_parameter="target_column",
                        options=[c.name for c in measure_cols],
                    )
                    return [], AgentResponseStatusEnum.CLARIFICATION, req, f"Pilih kolom numerik untuk perhitungan {agg_label}."
                elif table_entry.columns:
                    target_col = table_entry.columns[0]
                else:
                    return [], AgentResponseStatusEnum.UNSUPPORTED, None, "Tidak ditemukan kolom numerik untuk perhitungan."

        # 3. Determine Target Cell using Strict Priority (Rule 6, 7):
        # Priority 1: Explicit destination in prompt (e.g. "di D10", "in N20", "ke N20")
        # Priority 2: Selected cell in UI (e.g. "D10")
        # Priority 3: Safe automatic placement policy fallback
        if explicit_dest_cell:
            target_cell = explicit_dest_cell
        elif selected_range and ":" not in selected_range:
            target_cell = selected_range.upper()
        elif selected_range and selected_range.split(":")[0] == selected_range.split(":")[1]:
            target_cell = selected_range.split(":")[0].upper()
        else:
            placement = PlacementPolicy.determine_placement(
                table=table_entry,
                measure_col=target_col,
                grid=grid,
                query=user_request,
            )
            target_cell = placement.target_cell

        # 4. Formulate Formula & Calculate Python Truth
        if target_range_context and not requested_metric_keyword and not exact_matched_cols:
            formula_range = target_range_context
        else:
            col_letter = target_col.source_column_letter.upper()
            start_row = 2
            end_row = 101
            if table_entry.data_range and ":" in table_entry.data_range:
                try:
                    s_cell, e_cell = table_entry.data_range.split(":")
                    _, start_row = coordinate_from_string(s_cell)
                    _, end_row = coordinate_from_string(e_cell)
                except Exception:
                    pass
            formula_range = f"{col_letter}{start_row}:{col_letter}{end_row}"

        formula_str = f"={agg_func}({formula_range})"
        expected_val, _ = FormulaEvaluator.evaluate(formula_str, grid)
        is_replace_intent = any(kw in q_norm for kw in ["ganti", "replace", "overwrite", "timpa", "ubah isi"])

        # 5. State Inspection on Target Formula Cell
        try:
            dest_col_str, dest_row = coordinate_from_string(target_cell)
            dest_c_idx = column_index_from_string(dest_col_str)
            existing_cell = grid.get_cell(dest_row, dest_c_idx)
        except Exception as e:
            return [], AgentResponseStatusEnum.VALIDATION_ERROR, None, f"Koordinat sel '{target_cell}' tidak valid: {str(e)}"

        is_formula_already_satisfied = False
        if not existing_cell.is_empty:
            if existing_cell.formula:
                norm_existing = existing_cell.formula.strip().upper().replace(" ", "")
                norm_expected = formula_str.strip().upper().replace(" ", "")
                if norm_existing == norm_expected:
                    is_formula_already_satisfied = True
            elif existing_cell.parsed_value is not None and expected_val is not None:
                try:
                    if float(existing_cell.parsed_value) == float(expected_val):
                        is_formula_already_satisfied = True
                except Exception:
                    if str(existing_cell.parsed_value).strip() == str(expected_val).strip():
                        is_formula_already_satisfied = True

        if not is_formula_already_satisfied and not existing_cell.is_empty and existing_cell.parsed_value is not None and not is_replace_intent:
            req = ClarificationRequest(
                question=f"Target sel '{target_cell}' sudah terisi data ({existing_cell.formula or existing_cell.parsed_value}). Apakah ingin menimpa data tersebut?",
                reason=f"Target cell {target_cell} already contains conflicting data.",
                target_parameter="placement_collision",
                options=["Timpa data", "Batalkan"],
            )
            msg = (
                f"Target cell '{target_cell}' already contains another formula/value ({existing_cell.formula or existing_cell.parsed_value}). I won't overwrite it without explicit replace/overwrite intent."
                if is_english_query
                else f"Target sel '{target_cell}' sudah terisi data ({existing_cell.formula or existing_cell.parsed_value}). Saya tidak menimpanya tanpa instruksi untuk mengganti."
            )
            return [], AgentResponseStatusEnum.CLARIFICATION, req, msg

        # 6. Multi-Intent Check: Explicit Label Cell & Text
        is_summary_table_requested = any(kw in q_norm for kw in [
            "tabel rekap", "tabel summary", "rekapitulasi", "tulis label", "dengan label", "buatkan rekap"
        ])

        actions: List[SpreadsheetAction] = []
        action_idx = 1

        if explicit_label_cell and explicit_label_text:
            try:
                lbl_col_str, lbl_row = coordinate_from_string(explicit_label_cell)
                lbl_c_idx = column_index_from_string(lbl_col_str)
                existing_label_cell = grid.get_cell(lbl_row, lbl_c_idx)
            except Exception as e:
                return [], AgentResponseStatusEnum.VALIDATION_ERROR, None, f"Koordinat label '{explicit_label_cell}' tidak valid: {str(e)}"

            is_label_already_satisfied = False
            if not existing_label_cell.is_empty:
                if str(existing_label_cell.parsed_value).strip().lower() == explicit_label_text.strip().lower() or str(existing_label_cell.original_value).strip().lower() == explicit_label_text.strip().lower():
                    is_label_already_satisfied = True

            if not is_label_already_satisfied and not existing_label_cell.is_empty and existing_label_cell.parsed_value is not None and not is_replace_intent:
                req = ClarificationRequest(
                    question=f"Target sel label '{explicit_label_cell}' sudah terisi data ({existing_label_cell.parsed_value}). Apakah ingin menimpa data tersebut?",
                    reason=f"Label cell {explicit_label_cell} already contains conflicting data.",
                    target_parameter="placement_collision",
                    options=["Timpa data", "Batalkan"],
                )
                msg = (
                    f"Label cell '{explicit_label_cell}' already contains data ({existing_label_cell.parsed_value}). I won't overwrite it without explicit replace/overwrite intent."
                    if is_english_query
                    else f"Target sel label '{explicit_label_cell}' sudah terisi data ({existing_label_cell.parsed_value}). Saya tidak menimpanya tanpa instruksi untuk mengganti."
                )
                return [], AgentResponseStatusEnum.CLARIFICATION, req, msg

            if not is_label_already_satisfied:
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{action_idx}",
                        action_type=ActionTypeEnum.WRITE_VALUE,
                        sheet_name=cur_sheet,
                        target_cell=explicit_label_cell,
                        value=explicit_label_text,
                        description=f"Write label '{explicit_label_text}' to cell {explicit_label_cell}",
                    )
                )
                action_idx += 1

            if not is_formula_already_satisfied:
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{action_idx}",
                        action_type=ActionTypeEnum.WRITE_FORMULA,
                        sheet_name=cur_sheet,
                        target_cell=target_cell,
                        formula=formula_str,
                        expected_result=expected_val,
                        description=f"Write formula '{formula_str}' to cell {target_cell}",
                    )
                )
                action_idx += 1

            if explicit_format_type == "currency":
                fmt_target = explicit_format_cell or target_cell
                fmt_str = "Rp#,##0" if any(kw in q_norm for kw in ["rupiah", "rp", "idr"]) else "$#,##0.00"
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{action_idx}",
                        action_type=ActionTypeEnum.SET_NUMBER_FORMAT,
                        sheet_name=cur_sheet,
                        target_cell=fmt_target,
                        number_format=fmt_str,
                        description=f"Apply currency format '{fmt_str}' to {fmt_target}",
                    )
                )
                action_idx += 1
            elif explicit_format_type == "percentage":
                fmt_target = explicit_format_cell or target_cell
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{action_idx}",
                        action_type=ActionTypeEnum.SET_NUMBER_FORMAT,
                        sheet_name=cur_sheet,
                        target_cell=fmt_target,
                        number_format="0.0%",
                        description=f"Apply percentage format '0.0%' to {fmt_target}",
                    )
                )
                action_idx += 1
            elif explicit_format_type == "bold":
                fmt_target = explicit_format_cell or target_cell
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{action_idx}",
                        action_type=ActionTypeEnum.FORMAT_CELL,
                        sheet_name=cur_sheet,
                        target_cell=fmt_target,
                        style=FormattingStyle(bold=True),
                        description=f"Apply bold formatting to {fmt_target}",
                    )
                )
                action_idx += 1

            if actions:
                val_res = ActionValidator.validate_sequence(actions, workbook_index, grid)
                if not val_res.is_valid:
                    return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."

            if is_formula_already_satisfied and is_label_already_satisfied:
                confirm_msg = (
                    f"{agg_label} {target_col.name} in {target_cell} and label '{explicit_label_text}' in {explicit_label_cell} already exist, so no changes were made."
                    if is_english_query
                    else f"{agg_label} {target_col.name} di {target_cell} dan label '{explicit_label_text}' di {explicit_label_cell} sudah sesuai, jadi tidak ada perubahan yang dibuat."
                )
            elif is_formula_already_satisfied and not is_label_already_satisfied:
                confirm_msg = (
                    f"{agg_label} {target_col.name} in {target_cell} is already correct. Label '{explicit_label_text}' was added to {explicit_label_cell}."
                    if is_english_query
                    else f"{agg_label} {target_col.name} di {target_cell} sudah benar. Label '{explicit_label_text}' ditambahkan di {explicit_label_cell}."
                )
            elif not is_formula_already_satisfied and is_label_already_satisfied:
                confirm_msg = (
                    f"Label '{explicit_label_text}' in {explicit_label_cell} already exists. {agg_label} {target_col.name} was calculated in {target_cell}."
                    if is_english_query
                    else f"Label '{explicit_label_text}' di {explicit_label_cell} sudah ada. {agg_label} {target_col.name} dihitung di {target_cell}."
                )
            else:
                confirm_msg = (
                    f"Done. {agg_label} {target_col.name} was calculated in {target_cell} with label '{explicit_label_text}' in {explicit_label_cell}."
                    if is_english_query
                    else f"Selesai. {agg_label} {target_col.name} dihitung di {target_cell} dan label '{explicit_label_text}' ditulis di {explicit_label_cell}."
                )

            return actions, AgentResponseStatusEnum.SUCCESS, None, confirm_msg

        if is_summary_table_requested:
            # Multi-cell summary table explicitly requested
            col_s, row_i = coordinate_from_string(target_cell)
            col_i = column_index_from_string(col_s)
            label_cell = f"{get_column_letter(max(1, col_i - 1))}{row_i}" if col_i > 1 else None
            if label_cell:
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{action_idx}",
                        action_type=ActionTypeEnum.WRITE_VALUE,
                        sheet_name=cur_sheet,
                        target_cell=label_cell,
                        value=f"{agg_label} {target_col.name}",
                        description=f"Write label '{agg_label} {target_col.name}'",
                    )
                )
                action_idx += 1

            actions.append(
                SpreadsheetAction(
                    action_id=f"act_{action_idx}",
                    action_type=ActionTypeEnum.WRITE_FORMULA,
                    sheet_name=cur_sheet,
                    target_cell=target_cell,
                    formula=formula_str,
                    expected_result=expected_val,
                    description=f"Write formula '{formula_str}'",
                )
            )
            action_idx += 1

            if label_cell:
                actions.append(
                    SpreadsheetAction(
                        action_id=f"act_{action_idx}",
                        action_type=ActionTypeEnum.FORMAT_RANGE,
                        sheet_name=cur_sheet,
                        target_range=f"{label_cell}:{target_cell}",
                        style=FormattingStyle(bold=True, fill_color="#F1F5F9"),
                        description="Format summary row",
                    )
                )
                action_idx += 1
        else:
            # Minimal single formula write (Idempotent check)
            if is_formula_already_satisfied:
                confirm_msg = (
                    f"{agg_label} {target_col.name} is already calculated in {target_cell}, so no changes were made."
                    if is_english_query
                    else f"{agg_label} {target_col.name} sudah tersedia di {target_cell}, jadi tidak ada perubahan yang dibuat."
                )
                return [], AgentResponseStatusEnum.SUCCESS, None, confirm_msg

            actions.append(
                SpreadsheetAction(
                    action_id=f"act_{action_idx}",
                    action_type=ActionTypeEnum.WRITE_FORMULA,
                    sheet_name=cur_sheet,
                    target_cell=target_cell,
                    formula=formula_str,
                    expected_result=expected_val,
                    description=f"Write formula '{formula_str}' to cell {target_cell}",
                )
            )
            action_idx += 1

        # 7. Validate Action Sequence
        val_res = ActionValidator.validate_sequence(actions, workbook_index, grid)
        if not val_res.is_valid:
            return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."

        # Response Message reflecting exact operation and language parity (Rule 19, 20)
        if is_english_query:
            confirm_msg = f"Done. {agg_label} {target_col.name} was calculated in {target_cell}."
        else:
            confirm_msg = f"Selesai. {agg_label} {target_col.name} dihitung di {target_cell}."

        return actions, AgentResponseStatusEnum.SUCCESS, None, confirm_msg

    @classmethod
    def _plan_visualization_actions(
        cls,
        user_request: str,
        workbook_index: WorkbookMetadataIndex,
        grid: RawSheetGrid,
        cur_sheet: str,
        table_entry: TableIndexEntry,
        selected_range: Optional[str] = None,
        explicit_dest_cell: Optional[str] = None,
        target_range_context: Optional[str] = None,
    ) -> Optional[Tuple[List[SpreadsheetAction], AgentResponseStatusEnum, Optional[ClarificationRequest], str]]:
        """Handles visualization, chart modification, dashboard generation, and KPI actions."""
        q_norm = user_request.strip().lower()

        # 1. MOVE CHART
        if (any(kw in q_norm for kw in ["pindahkan", "move", "geser", "relocate"]) and any(c in q_norm for c in ["chart", "grafik", "diagram", "plot"])) or any(kw in q_norm for kw in ["pindahkan ke", "move to"]):
            move_target = explicit_dest_cell or (selected_range.split(":")[0].upper() if selected_range else None) or "H20"
            chart_id = list(grid.charts.keys())[0] if grid.charts else f"chart_{uuid.uuid4().hex[:8]}"
            action = SpreadsheetAction(
                action_id="act_1",
                action_type=ActionTypeEnum.MOVE_CHART,
                sheet_name=cur_sheet,
                target_cell=move_target,
                chart_spec=ChartActionSpec(
                    chart_id=chart_id,
                    sheet_name=cur_sheet,
                    chart_type="BAR",
                    title="Moved Chart",
                    destination_cell=move_target,
                    anchor_cell=move_target,
                ),
                description=f"Move chart to {move_target}",
            )
            val_res = ActionValidator.validate_sequence([action], workbook_index, grid)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."
            return [action], AgentResponseStatusEnum.SUCCESS, None, f"Chart berhasil dipindahkan ke {move_target}."

        # 2. RESIZE CHART
        if any(kw in q_norm for kw in ["resize chart", "ubah ukuran chart", "perbesar chart", "perkecil chart"]):
            w = 10
            h = 20
            w_match = re.search(r'(\d+)\s*(?:kolom|cols|columns)', q_norm)
            h_match = re.search(r'(\d+)\s*(?:baris|rows)', q_norm)
            if w_match:
                w = int(w_match.group(1))
            if h_match:
                h = int(h_match.group(1))
            dest = explicit_dest_cell or (selected_range.split(":")[0].upper() if selected_range else None) or (list(grid.charts.values())[0].get("destination_cell") if grid.charts else cls._resolve_default_chart_anchor(table_entry))
            chart_id = list(grid.charts.keys())[0] if grid.charts else f"chart_{uuid.uuid4().hex[:8]}"
            action = SpreadsheetAction(
                action_id="act_1",
                action_type=ActionTypeEnum.RESIZE_CHART,
                sheet_name=cur_sheet,
                target_cell=dest,
                chart_spec=ChartActionSpec(
                    chart_id=chart_id,
                    sheet_name=cur_sheet,
                    chart_type="BAR",
                    title="Resized Chart",
                    destination_cell=dest,
                    width_cols=w,
                    height_rows=h,
                ),
                description=f"Resize chart at {dest} to {w} cols x {h} rows",
            )
            val_res = ActionValidator.validate_sequence([action], workbook_index, grid)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."
            return [action], AgentResponseStatusEnum.SUCCESS, None, f"Ukuran chart di {dest} berhasil diubah menjadi {w} kolom x {h} baris."

        # 3. DELETE CHART
        if any(kw in q_norm for kw in ["hapus chart", "delete chart", "remove chart", "buang chart", "hilangkan chart"]):
            del_target = explicit_dest_cell or (selected_range.split(":")[0].upper() if selected_range else None) or (list(grid.charts.values())[0].get("destination_cell") if grid.charts else cls._resolve_default_chart_anchor(table_entry))
            action = SpreadsheetAction(
                action_id="act_1",
                action_type=ActionTypeEnum.DELETE_CHART,
                sheet_name=cur_sheet,
                target_cell=del_target,
                description=f"Delete chart at {del_target}",
            )
            val_res = ActionValidator.validate_sequence([action], workbook_index, grid)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."
            return [action], AgentResponseStatusEnum.SUCCESS, None, f"Chart di {del_target} berhasil dihapus."

        # 4. UPDATE CHART
        if any(kw in q_norm for kw in ["ubah chart", "ganti chart", "update chart", "change chart"]):
            new_type_str = "BAR"
            if "pie" in q_norm:
                new_type_str = "PIE"
            elif "line" in q_norm:
                new_type_str = "LINE"
            elif "area" in q_norm:
                new_type_str = "AREA"
            elif "scatter" in q_norm:
                new_type_str = "SCATTER"
            elif "column" in q_norm:
                new_type_str = "COLUMN"

            upd_target = explicit_dest_cell or (selected_range.split(":")[0].upper() if selected_range else None) or (list(grid.charts.values())[0].get("destination_cell") if grid.charts else cls._resolve_default_chart_anchor(table_entry))
            chart_spec = ChartActionSpec(
                chart_id=f"chart_{uuid.uuid4().hex[:8]}",
                sheet_name=cur_sheet,
                chart_type=new_type_str,
                title=f"Updated Chart ({new_type_str})",
                destination_cell=upd_target,
                anchor_cell=upd_target,
            )
            action = SpreadsheetAction(
                action_id="act_1",
                action_type=ActionTypeEnum.UPDATE_CHART,
                sheet_name=cur_sheet,
                target_cell=upd_target,
                chart_spec=chart_spec,
                description=f"Update chart at {upd_target} to {new_type_str}",
            )
            val_res = ActionValidator.validate_sequence([action], workbook_index, grid)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."
            return [action], AgentResponseStatusEnum.SUCCESS, None, f"Chart di {upd_target} berhasil diubah menjadi {new_type_str} chart."

        # 5. CREATE KPI
        is_kpi_req = ("kpi" in q_norm or "kartu kpi" in q_norm) and not any(kw in q_norm for kw in ["dashboard", "dasbor"])
        if is_kpi_req:
            measure_cols = [c for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.INTEGER, DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.PERCENTAGE}]
            target_m_col = measure_cols[0] if measure_cols else table_entry.columns[-1]
            for c in table_entry.columns:
                if c.normalized_name in q_norm or c.name.lower() in q_norm:
                    target_m_col = c
                    break

            agg_name = "SUM"
            if any(kw in q_norm for kw in ["rata-rata", "average", "mean"]):
                agg_name = "AVERAGE"
            elif any(kw in q_norm for kw in ["count", "jumlah", "banyaknya"]):
                agg_name = "COUNT_ROWS"
            elif "min" in q_norm:
                agg_name = "MIN"
            elif "max" in q_norm:
                agg_name = "MAX"

            col_idx = target_m_col.index + 1
            col_vals = []
            for r in range(2, min(grid.max_row + 1, 1002)):
                cell = grid.get_cell(r, col_idx)
                if cell and not cell.is_empty and cell.parsed_value is not None:
                    col_vals.append(cell.parsed_value)

            s_series = pd.Series(col_vals)
            raw_val, formatted_val, _ = DeterministicAggregator.calculate_scalar(
                s_series, OperationEnum(agg_name) if agg_name in OperationEnum._value2member_map_ else OperationEnum.SUM, len(col_vals)
            )

            dest_c = explicit_dest_cell or "G2"
            kpi_spec = KPIActionSpec(
                kpi_id=f"kpi_{uuid.uuid4().hex[:8]}",
                title=f"{agg_name} {target_m_col.name}",
                measure_column=target_m_col.name,
                aggregation=agg_name,
                calculated_value=raw_val or 0.0,
                formatted_value=formatted_val or str(raw_val),
                destination_cell=dest_c,
                source_range=table_entry.range_address,
            )
            action = SpreadsheetAction(
                action_id="act_1",
                action_type=ActionTypeEnum.CREATE_KPI,
                sheet_name=cur_sheet,
                target_cell=dest_c,
                kpi_spec=kpi_spec,
                description=f"Create KPI '{kpi_spec.title}' at {dest_c}",
            )
            val_res = ActionValidator.validate_sequence([action], workbook_index, grid)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."
            return [action], AgentResponseStatusEnum.SUCCESS, None, f"KPI '{kpi_spec.title}' berhasil dibuat di {dest_c} (Nilai: {formatted_val})."

        # 6. MULTI-VISUALIZATION & DASHBOARD GENERATION
        is_multi_vis = any(kw in q_norm for kw in [
            "semua kemungkinan", "semua visualisasi", "semua chart", "semua grafik",
            "berikan semuanya", "tampilkan semuanya", "jangan hanya itu", "jangan cuma itu",
            "visualize everything", "all visualizations", "all useful visualizations", "all charts",
            "visualisasikan data ini", "visualize this data", "dashboard visualisasi", "buat semua chart",
            "buatkan semua chart", "semua visualisasi yang relevan", "semua chart yang cocok",
        ])
        if is_multi_vis:
            return cls._plan_multi_visualizations(
                grid=grid,
                workbook_index=workbook_index,
                cur_sheet=cur_sheet,
                table_entry=table_entry,
                q_norm=q_norm,
                explicit_dest_cell=explicit_dest_cell,
                selected_range=selected_range,
                target_range_context=target_range_context,
            )

        is_dash_req = any(kw in q_norm for kw in [
            "dashboard", "dasbor", "buat data ini menjadi dashboard", "executive dashboard", "siap dishare ke atasan", "siap dipresentasikan", "ringkasan visual", "visual summary"
        ])
        if is_dash_req:
            actions: List[SpreadsheetAction] = []
            act_idx = 1
            is_dedicated = any(kw in q_norm for kw in [
                "buat data ini menjadi dashboard", "executive dashboard", "siap dishare", "dipresentasikan", "sheet dashboard", "lembar dashboard", "ke atasan"
            ])
            dash_sheet = "Dashboard" if is_dedicated else cur_sheet

            num_cols = [
                c for c in table_entry.columns
                if (c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.PERCENTAGE})
                and c.semantic_type != SemanticTypeEnum.IDENTIFIER
                and not c.name.lower().endswith("id")
                and not c.name.lower().startswith("id")
            ]
            if not num_cols:
                num_cols = [c for c in table_entry.columns if c.data_type in {DataTypeEnum.INTEGER, DataTypeEnum.FLOAT}]

            cat_cols = [c for c in table_entry.columns if (c.semantic_type == SemanticTypeEnum.CATEGORICAL or c.data_type == DataTypeEnum.STRING) and not c.name.lower().endswith("id")]
            if not cat_cols:
                cat_cols = [c for c in table_entry.columns if c.data_type == DataTypeEnum.STRING]

            temp_cols = [c for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.TEMPORAL or c.data_type in {DataTypeEnum.DATE, DataTypeEnum.DATETIME}]

            m_primary = num_cols[0] if num_cols else table_entry.columns[-1]
            m_secondary = num_cols[1] if len(num_cols) > 1 else m_primary
            dim_primary = cat_cols[0] if cat_cols else (table_entry.columns[1] if len(table_entry.columns) > 1 else table_entry.columns[0])
            dim_temp = temp_cols[0] if temp_cols else (cat_cols[1] if len(cat_cols) > 1 else dim_primary)

            if is_dedicated:
                # 1. Create Dashboard worksheet
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.CREATE_WORKSHEET,
                    sheet_name=dash_sheet,
                    description=f"Create dedicated '{dash_sheet}' worksheet",
                ))
                act_idx += 1

                # 2. Executive Title Header
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.WRITE_VALUE,
                    sheet_name=dash_sheet,
                    target_cell="A1",
                    value="SALES PERFORMANCE DASHBOARD",
                    style=FormattingStyle(bold=True, font_size=14, fill_color="#1E293B", font_color="#FFFFFF"),
                    description="Write Dashboard Title Header",
                ))
                act_idx += 1

                # 3. Subtitle Header
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.WRITE_VALUE,
                    sheet_name=dash_sheet,
                    target_cell="A2",
                    value="Executive overview of performance metrics, distribution, and trends",
                    style=FormattingStyle(italic=True, font_size=10, font_color="#64748B"),
                    description="Write Dashboard Subtitle",
                ))
                act_idx += 1

                # KPI 1: Primary Measure at A4:A5
                kpi1_vals = [grid.get_cell(r, m_primary.index + 1).parsed_value for r in range(2, min(grid.max_row + 1, 1002)) if grid.get_cell(r, m_primary.index + 1).parsed_value is not None]
                raw1, fmt1, _ = DeterministicAggregator.calculate_scalar(pd.Series(kpi1_vals), OperationEnum.SUM, len(kpi1_vals))
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.WRITE_VALUE,
                    sheet_name=dash_sheet,
                    target_cell="A4",
                    value=f"Total {m_primary.name}",
                    style=FormattingStyle(bold=True, font_size=9, font_color="#64748B"),
                    description="Write KPI 1 label",
                ))
                act_idx += 1
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.WRITE_VALUE,
                    sheet_name=dash_sheet,
                    target_cell="A5",
                    value=raw1,
                    number_format="$#,##0.00" if m_primary.data_type == DataTypeEnum.CURRENCY else "#,##0.00",
                    style=FormattingStyle(bold=True, font_size=12, fill_color="#F8FAFC"),
                    description="Write KPI 1 value",
                ))
                act_idx += 1
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.CREATE_KPI,
                    sheet_name=dash_sheet,
                    target_cell="A4",
                    kpi_spec=KPIActionSpec(
                        kpi_id=f"kpi_{uuid.uuid4().hex[:8]}",
                        title=f"Total {m_primary.name}",
                        measure_column=m_primary.name,
                        aggregation="SUM",
                        calculated_value=raw1 or 0.0,
                        formatted_value=fmt1 or str(raw1),
                        destination_cell="A4",
                        source_range=table_entry.range_address,
                    ),
                    description="Register Dashboard KPI 1",
                ))
                act_idx += 1

                # KPI 2: Total Orders / Row Count at C4:C5
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.WRITE_VALUE,
                    sheet_name=dash_sheet,
                    target_cell="C4",
                    value="Total Records",
                    style=FormattingStyle(bold=True, font_size=9, font_color="#64748B"),
                    description="Write KPI 2 label",
                ))
                act_idx += 1
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.WRITE_VALUE,
                    sheet_name=dash_sheet,
                    target_cell="C5",
                    value=len(kpi1_vals),
                    number_format="#,##0",
                    style=FormattingStyle(bold=True, font_size=12, fill_color="#F8FAFC"),
                    description="Write KPI 2 value",
                ))
                act_idx += 1
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.CREATE_KPI,
                    sheet_name=dash_sheet,
                    target_cell="C4",
                    kpi_spec=KPIActionSpec(
                        kpi_id=f"kpi_{uuid.uuid4().hex[:8]}",
                        title="Total Records",
                        measure_column="OrderID",
                        aggregation="COUNT",
                        calculated_value=len(kpi1_vals),
                        formatted_value=f"{len(kpi1_vals):,}",
                        destination_cell="C4",
                        source_range=table_entry.range_address,
                    ),
                    description="Register Dashboard KPI 2",
                ))
                act_idx += 1

                # KPI 3: Average Primary Metric at E4:E5
                raw_avg, fmt_avg, _ = DeterministicAggregator.calculate_scalar(pd.Series(kpi1_vals), OperationEnum.AVERAGE, len(kpi1_vals))
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.WRITE_VALUE,
                    sheet_name=dash_sheet,
                    target_cell="E4",
                    value=f"Average {m_primary.name}",
                    style=FormattingStyle(bold=True, font_size=9, font_color="#64748B"),
                    description="Write KPI 3 label",
                ))
                act_idx += 1
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.WRITE_VALUE,
                    sheet_name=dash_sheet,
                    target_cell="E5",
                    value=raw_avg,
                    number_format="$#,##0.00" if m_primary.data_type == DataTypeEnum.CURRENCY else "#,##0.00",
                    style=FormattingStyle(bold=True, font_size=12, fill_color="#F8FAFC"),
                    description="Write KPI 3 value",
                ))
                act_idx += 1
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.CREATE_KPI,
                    sheet_name=dash_sheet,
                    target_cell="E4",
                    kpi_spec=KPIActionSpec(
                        kpi_id=f"kpi_{uuid.uuid4().hex[:8]}",
                        title=f"Average {m_primary.name}",
                        measure_column=m_primary.name,
                        aggregation="AVERAGE",
                        calculated_value=raw_avg or 0.0,
                        formatted_value=fmt_avg or str(raw_avg),
                        destination_cell="E4",
                        source_range=table_entry.range_address,
                    ),
                    description="Register Dashboard KPI 3",
                ))
                act_idx += 1

                # Chart 1: PIE / BAR at A8
                c1_type = ChartTypeEnum.PIE if len(cat_cols) > 0 else ChartTypeEnum.BAR
                c1_spec = cls._build_chart_spec(
                    grid=grid,
                    workbook_index=workbook_index,
                    cur_sheet=dash_sheet,
                    table_entry=table_entry,
                    dim_col=dim_primary,
                    measure_col=m_primary,
                    chart_type=c1_type,
                    agg_name="SUM",
                    dest_cell="A8",
                    width_cols=6,
                    height_rows=14,
                )
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.CREATE_CHART,
                    sheet_name=dash_sheet,
                    target_cell="A8",
                    chart_spec=c1_spec,
                    description=f"Create Dashboard {c1_type.value} chart at A8",
                ))
                act_idx += 1

                # Chart 2: LINE / BAR at E8
                c2_type = ChartTypeEnum.LINE if temp_cols else ChartTypeEnum.BAR
                c2_spec = cls._build_chart_spec(
                    grid=grid,
                    workbook_index=workbook_index,
                    cur_sheet=dash_sheet,
                    table_entry=table_entry,
                    dim_col=dim_temp,
                    measure_col=m_primary,
                    chart_type=c2_type,
                    agg_name="SUM",
                    dest_cell="E8",
                    width_cols=6,
                    height_rows=14,
                )
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.CREATE_CHART,
                    sheet_name=dash_sheet,
                    target_cell="E8",
                    chart_spec=c2_spec,
                    description=f"Create Dashboard {c2_type.value} chart at E8",
                ))
                act_idx += 1
            else:
                # In-sheet visual dashboard at G2:R17
                kpi1_vals = [grid.get_cell(r, m_primary.index + 1).parsed_value for r in range(2, min(grid.max_row + 1, 1002)) if grid.get_cell(r, m_primary.index + 1).parsed_value is not None]
                raw1, fmt1, _ = DeterministicAggregator.calculate_scalar(pd.Series(kpi1_vals), OperationEnum.SUM, len(kpi1_vals))
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.CREATE_KPI,
                    sheet_name=cur_sheet,
                    target_cell="G2",
                    kpi_spec=KPIActionSpec(
                        kpi_id=f"kpi_{uuid.uuid4().hex[:8]}",
                        title=f"Total {m_primary.name}",
                        measure_column=m_primary.name,
                        aggregation="SUM",
                        calculated_value=raw1 or 0.0,
                        formatted_value=fmt1 or str(raw1),
                        destination_cell="G2",
                        source_range=table_entry.range_address,
                    ),
                    description=f"Create Dashboard KPI 'Total {m_primary.name}' at G2",
                ))
                act_idx += 1

                kpi2_vals = [grid.get_cell(r, m_secondary.index + 1).parsed_value for r in range(2, min(grid.max_row + 1, 1002)) if grid.get_cell(r, m_secondary.index + 1).parsed_value is not None]
                raw2, fmt2, _ = DeterministicAggregator.calculate_scalar(pd.Series(kpi2_vals), OperationEnum.SUM if len(num_cols) > 1 else OperationEnum.AVERAGE, len(kpi2_vals))
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.CREATE_KPI,
                    sheet_name=cur_sheet,
                    target_cell="J2",
                    kpi_spec=KPIActionSpec(
                        kpi_id=f"kpi_{uuid.uuid4().hex[:8]}",
                        title=f"Total {m_secondary.name}" if len(num_cols) > 1 else f"Avg {m_primary.name}",
                        measure_column=m_secondary.name,
                        aggregation="SUM" if len(num_cols) > 1 else "AVERAGE",
                        calculated_value=raw2 or 0.0,
                        formatted_value=fmt2 or str(raw2),
                        destination_cell="J2",
                        source_range=table_entry.range_address,
                    ),
                    description="Create Dashboard KPI at J2",
                ))
                act_idx += 1

                c1_type = ChartTypeEnum.PIE if len(cat_cols) > 0 else ChartTypeEnum.BAR
                c1_spec = cls._build_chart_spec(
                    grid=grid,
                    workbook_index=workbook_index,
                    cur_sheet=cur_sheet,
                    table_entry=table_entry,
                    dim_col=dim_primary,
                    measure_col=m_primary,
                    chart_type=c1_type,
                    agg_name="SUM",
                    dest_cell="G5",
                    width_cols=6,
                    height_rows=12,
                )
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.CREATE_CHART,
                    sheet_name=cur_sheet,
                    target_cell="G5",
                    chart_spec=c1_spec,
                    description=f"Create Dashboard {c1_type.value} chart at G5",
                ))
                act_idx += 1

                c2_type = ChartTypeEnum.LINE if temp_cols else ChartTypeEnum.BAR
                c2_spec = cls._build_chart_spec(
                    grid=grid,
                    workbook_index=workbook_index,
                    cur_sheet=cur_sheet,
                    table_entry=table_entry,
                    dim_col=dim_temp,
                    measure_col=m_primary,
                    chart_type=c2_type,
                    agg_name="SUM",
                    dest_cell="M5",
                    width_cols=6,
                    height_rows=12,
                )
                actions.append(SpreadsheetAction(
                    action_id=f"act_{act_idx}",
                    action_type=ActionTypeEnum.CREATE_CHART,
                    sheet_name=cur_sheet,
                    target_cell="M5",
                    chart_spec=c2_spec,
                    description=f"Create Dashboard {c2_type.value} chart at M5",
                ))
                act_idx += 1

            val_res = ActionValidator.validate_sequence(actions, workbook_index, grid)
            if not val_res.is_valid:
                return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."
            return actions, AgentResponseStatusEnum.SUCCESS, None, f"Dashboard {m_primary.name} berhasil dibuat."

        # 5. SINGLE CREATE CHART
        dim_col, measure_col = cls._resolve_chart_columns(
            q_norm=q_norm,
            table_entry=table_entry,
            target_range_context=target_range_context,
        )

        chart_type = cls._resolve_chart_type(q_norm=q_norm, dim_col=dim_col, measure_col=measure_col)

        agg_name = "SUM"
        if any(kw in q_norm for kw in ["rata-rata", "average", "mean"]):
            agg_name = "AVERAGE"
        elif any(kw in q_norm for kw in ["count", "jumlah", "banyaknya", "banyak"]):
            agg_name = "COUNT"
        elif "min" in q_norm:
            agg_name = "MIN"
        elif "max" in q_norm:
            agg_name = "MAX"

        top_n = None
        top_match = re.search(r'\b(?:top|tertinggi|teratas)\s+(\d+)\b', q_norm) or re.search(r'\b(\d+)\s+(?:tertinggi|teratas|product|produk|customer)\b', q_norm)
        if top_match:
            top_n = int(top_match.group(1))

        is_english_query = any(w in q_norm.split() for w in ["pie", "bar", "chart", "sales", "profit", "create", "make", "show", "generate"]) and not any(w in q_norm.split() for w in ["buatkan", "buat", "tampilkan", "grafik", "wilayah", "kategori"])

        # Chart Idempotency Check: Verify if equivalent chart already exists
        if grid and grid.charts:
            for ex_id, ex_c in grid.charts.items():
                ex_dim = (ex_c.get("dimension_column") or "").strip().lower()
                ex_meas = (ex_c.get("measure_column") or "").strip().lower()
                ex_type = (ex_c.get("chart_type") or "").strip().lower()
                ex_agg = (ex_c.get("aggregation") or "sum").strip().lower()
                ex_dest = ex_c.get("destination_cell") or ex_c.get("anchor_cell") or "B12"

                if (
                    ex_dim == dim_col.name.strip().lower()
                    and ex_meas == measure_col.name.strip().lower()
                    and ex_type == chart_type.value.strip().lower()
                    and ex_agg == agg_name.strip().lower()
                ):
                    title_name = ex_c.get("title") or f"{measure_col.name} by {dim_col.name}"
                    msg = (
                        f'{chart_type.value.capitalize()} chart "{title_name}" already exists at {ex_dest}, so no duplicate was created.'
                        if is_english_query
                        else f'{chart_type.value.capitalize()} chart "{title_name}" sudah tersedia di {ex_dest}, jadi saya tidak membuat duplikat.'
                    )
                    return [], AgentResponseStatusEnum.SUCCESS, None, msg

        # Resolve Destination Cell
        dest_cell = explicit_dest_cell or (selected_range.split(":")[0].upper() if selected_range else None) or cls._resolve_default_chart_anchor(table_entry, grid)

        # Spatial Collision Check on Explicit Destination
        if explicit_dest_cell and grid and grid.charts:
            try:
                dest_col_str, dest_row = coordinate_from_string(dest_cell)
                dest_col = column_index_from_string(dest_col_str)
                w = 7
                h = 14
                new_box = (dest_col, dest_row, dest_col + w - 1, dest_row + h - 1)
                for ex_id, ex_c in grid.charts.items():
                    ex_dest = ex_c.get("destination_cell") or ex_c.get("anchor_cell")
                    if not ex_dest:
                        continue
                    ex_col_str, ex_row = coordinate_from_string(ex_dest.strip().upper())
                    ex_col = column_index_from_string(ex_col_str)
                    ex_w = ex_c.get("width_cols", 7)
                    ex_h = ex_c.get("height_rows", 14)
                    ex_box = (ex_col, ex_row, ex_col + ex_w - 1, ex_row + ex_h - 1)
                    if not (new_box[2] < ex_box[0] or new_box[0] > ex_box[2] or new_box[3] < ex_box[1] or new_box[1] > ex_box[3]):
                        chart_title = ex_c.get("title") or "Chart"
                        req = ClarificationRequest(
                            question=f"Target sel '{dest_cell}' tumpang tindih dengan chart '{chart_title}' di '{ex_dest}'. Silakan pilih sel tujuan lain yang kosong.",
                            reason="Chart spatial collision: target region overlaps existing chart.",
                            target_parameter="chart_placement_collision",
                            options=["Pilih lokasi lain", "Batalkan"],
                        )
                        err_msg = (
                            f"Target chart region '{dest_cell}' overlaps with existing chart '{chart_title}' at '{ex_dest}'."
                            if is_english_query
                            else f"Target sel '{dest_cell}' tumpang tindih dengan chart '{chart_title}' di '{ex_dest}'. Silakan pilih sel tujuan lain yang kosong."
                        )
                        return [], AgentResponseStatusEnum.CLARIFICATION, req, err_msg
            except Exception:
                pass

        chart_spec = cls._build_chart_spec(
            grid=grid,
            workbook_index=workbook_index,
            cur_sheet=cur_sheet,
            table_entry=table_entry,
            dim_col=dim_col,
            measure_col=measure_col,
            chart_type=chart_type,
            agg_name=agg_name,
            dest_cell=dest_cell,
            top_n=top_n,
            target_range_context=target_range_context,
        )

        action = SpreadsheetAction(
            action_id="act_1",
            action_type=ActionTypeEnum.CREATE_CHART,
            sheet_name=cur_sheet,
            target_cell=dest_cell,
            chart_spec=chart_spec,
            description=f"Create {chart_type.value} chart '{chart_spec.title}' at {dest_cell}",
        )

        val_res = ActionValidator.validate_sequence([action], workbook_index, grid)
        if not val_res.is_valid:
            return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."

        return [action], AgentResponseStatusEnum.SUCCESS, None, f"{chart_type.value.capitalize()} chart '{chart_spec.title}' berhasil dibuat di {dest_cell}."

    @classmethod
    def _resolve_chart_columns(
        cls,
        q_norm: str,
        table_entry: TableIndexEntry,
        target_range_context: Optional[str] = None,
    ) -> Tuple[ColumnIndexEntry, ColumnIndexEntry]:
        """Resolves the dimension (categorical/temporal) and measure (numeric) columns."""
        dim_col = None
        measure_col = None
        q_compact = q_norm.replace(" ", "").replace("_", "").replace("-", "")

        # Column Synonym Mapping for Indonesian / English
        synonyms = {
            "sales": ["sales", "penjualan", "omset", "omzet", "revenue", "pendapatan"],
            "profit": ["profit", "laba", "keuntungan", "margin"],
            "quantity": ["quantity", "qty", "kuantitas", "jumlah barang"],
            "discount": ["discount", "diskon", "potongan"],
            "region": ["region", "wilayah", "daerah", "area"],
            "category": ["category", "kategori", "jenis"],
            "subcategory": ["subcategory", "subkategori", "sub-kategori", "sub kategori"],
            "segment": ["segment", "segmen", "pasar"],
            "orderdate": ["orderdate", "order date", "tanggal order", "tgl order", "tanggal", "tgl", "waktu", "date", "bulan", "tahun"],
            "productname": ["productname", "product name", "nama produk", "produk", "product", "item"],
            "customername": ["customername", "customer name", "nama pelanggan", "pelanggan", "customer", "klien"],
        }

        # Check explicit columns or synonyms mentioned in query
        for c in table_entry.columns:
            c_norm = c.normalized_name
            c_compact = c.name.lower().replace(" ", "").replace("_", "").replace("-", "")

            matched = (
                c_norm in q_norm
                or c.name.lower() in q_norm
                or c_norm in q_compact
                or c_compact in q_compact
            )

            if not matched and c_norm in synonyms:
                matched = any(syn in q_norm or syn.replace(" ", "") in q_compact for syn in synonyms[c_norm])

            if matched:
                if c.semantic_type in {SemanticTypeEnum.CATEGORICAL, SemanticTypeEnum.TEMPORAL, SemanticTypeEnum.IDENTIFIER} or c.data_type in {DataTypeEnum.STRING, DataTypeEnum.DATE, DataTypeEnum.DATETIME}:
                    if not dim_col:
                        dim_col = c
                elif c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.INTEGER, DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.PERCENTAGE}:
                    if not measure_col:
                        measure_col = c

        # For Scatter plots: check if two numeric columns are requested
        is_scatter = any(kw in q_norm for kw in ["scatter", "sebaran", "korelasi"])
        if is_scatter:
            matched_num_cols = []
            for c in table_entry.columns:
                c_norm = c.normalized_name
                c_compact = c.name.lower().replace(" ", "").replace("_", "").replace("-", "")
                m = (c_norm in q_norm or c.name.lower() in q_norm or c_norm in q_compact or c_compact in q_compact)
                if not m and c_norm in synonyms:
                    m = any(syn in q_norm or syn.replace(" ", "") in q_compact for syn in synonyms[c_norm])
                if m and (c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.INTEGER, DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.PERCENTAGE}):
                    matched_num_cols.append(c)
            if len(matched_num_cols) >= 2:
                dim_col = matched_num_cols[0]
                measure_col = matched_num_cols[1]

        # Check target_range_context (e.g. B2:C21)
        if not dim_col or not measure_col:
            if target_range_context and ":" in target_range_context:
                try:
                    start_c, _ = target_range_context.split(":")
                    start_col_letter, _ = coordinate_from_string(start_c)
                    end_c = target_range_context.split(":")[1]
                    end_col_letter, _ = coordinate_from_string(end_c)
                    for c in table_entry.columns:
                        if c.source_column_letter.upper() == start_col_letter.upper():
                            dim_col = c
                        elif c.source_column_letter.upper() == end_col_letter.upper():
                            measure_col = c
                except Exception:
                    pass

        # Fallbacks
        if not dim_col:
            cat_cols = [c for c in table_entry.columns if c.semantic_type in {SemanticTypeEnum.CATEGORICAL, SemanticTypeEnum.TEMPORAL} or c.data_type in {DataTypeEnum.STRING, DataTypeEnum.DATE, DataTypeEnum.DATETIME}]
            dim_col = cat_cols[0] if cat_cols else table_entry.columns[0]

        if not measure_col:
            num_cols = [c for c in table_entry.columns if c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.INTEGER, DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.PERCENTAGE}]
            measure_col = num_cols[0] if num_cols else table_entry.columns[-1]

        return dim_col, measure_col

    @classmethod
    def _resolve_chart_type(
        cls,
        q_norm: str,
        dim_col: ColumnIndexEntry,
        measure_col: ColumnIndexEntry,
    ) -> ChartTypeEnum:
        """Deterministically selects chart type from explicit keywords or heuristics."""
        if "pie" in q_norm:
            return ChartTypeEnum.PIE
        if "bar" in q_norm:
            return ChartTypeEnum.BAR
        if "column" in q_norm:
            return ChartTypeEnum.COLUMN
        if "line" in q_norm or "garis" in q_norm or "tren" in q_norm or "trend" in q_norm:
            return ChartTypeEnum.LINE
        if "area" in q_norm:
            return ChartTypeEnum.AREA
        if "scatter" in q_norm:
            return ChartTypeEnum.SCATTER
        if "histogram" in q_norm:
            return ChartTypeEnum.HISTOGRAM

        # Heuristics
        if dim_col.semantic_type == SemanticTypeEnum.TEMPORAL or any(kw in dim_col.normalized_name for kw in ["date", "month", "year", "tanggal", "bulan", "tahun"]):
            return ChartTypeEnum.LINE

        if any(kw in q_norm for kw in ["share", "komposisi", "proporsi", "porsi", "persentase"]):
            return ChartTypeEnum.PIE

        if any(kw in q_norm for kw in ["top", "tertinggi", "terendah", "ranking"]):
            return ChartTypeEnum.COLUMN

        return ChartTypeEnum.BAR

    @classmethod
    def _build_chart_spec(
        cls,
        grid: RawSheetGrid,
        workbook_index: WorkbookMetadataIndex,
        cur_sheet: str,
        table_entry: TableIndexEntry,
        dim_col: ColumnIndexEntry,
        measure_col: ColumnIndexEntry,
        chart_type: ChartTypeEnum,
        agg_name: str,
        dest_cell: str,
        top_n: Optional[int] = None,
        target_range_context: Optional[str] = None,
        width_cols: int = 8,
        height_rows: int = 15,
    ) -> ChartActionSpec:
        """Gathers data from grid, calculates deterministic aggregates, and renders chart artifact."""
        start_row = 2
        end_row = min(grid.max_row, 1002)

        if target_range_context and ":" in target_range_context:
            try:
                start_c, end_c = target_range_context.split(":")
                _, start_row = coordinate_from_string(start_c)
                _, end_row = coordinate_from_string(end_c)
            except Exception:
                pass
        elif table_entry.data_range and ":" in table_entry.data_range:
            try:
                start_c, end_c = table_entry.data_range.split(":")
                _, start_row = coordinate_from_string(start_c)
                _, end_row = coordinate_from_string(end_c)
            except Exception:
                pass

        dim_idx = dim_col.index + 1
        m_idx = measure_col.index + 1

        grouped: Dict[str, List[float]] = {}
        for r in range(start_row, end_row + 1):
            d_cell = grid.get_cell(r, dim_idx)
            m_cell = grid.get_cell(r, m_idx)
            d_val = str(d_cell.parsed_value) if (d_cell and not d_cell.is_empty and d_cell.parsed_value is not None) else "Unknown"

            m_val = None
            if m_cell and not m_cell.is_empty and m_cell.parsed_value is not None:
                try:
                    m_val = float(m_cell.parsed_value)
                except (ValueError, TypeError):
                    pass

            if d_val not in grouped:
                grouped[d_val] = []
            if m_val is not None:
                grouped[d_val].append(m_val)

        summary_dict: Dict[str, float] = {}
        for d, vals in grouped.items():
            if agg_name == "AVERAGE":
                summary_dict[d] = sum(vals) / len(vals) if vals else 0.0
            elif agg_name == "COUNT":
                summary_dict[d] = float(len(vals))
            elif agg_name == "MIN":
                summary_dict[d] = min(vals) if vals else 0.0
            elif agg_name == "MAX":
                summary_dict[d] = max(vals) if vals else 0.0
            else:
                summary_dict[d] = sum(vals)

        if chart_type == ChartTypeEnum.SCATTER and (dim_col.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or dim_col.data_type in {DataTypeEnum.INTEGER, DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY}):
            x_cats = []
            y_vals = []
            for r in range(start_row, end_row + 1):
                d_cell = grid.get_cell(r, dim_idx)
                m_cell = grid.get_cell(r, m_idx)
                if d_cell and not d_cell.is_empty and d_cell.parsed_value is not None and m_cell and not m_cell.is_empty and m_cell.parsed_value is not None:
                    try:
                        x_cats.append(str(float(d_cell.parsed_value)))
                        y_vals.append(float(m_cell.parsed_value))
                    except (ValueError, TypeError):
                        pass
            title = f"{measure_col.name} vs {dim_col.name}"
            summary_dict = {f"Point_{i+1}": y for i, y in enumerate(y_vals)}
        elif top_n:
            sorted_items = sorted(summary_dict.items(), key=lambda x: x[1], reverse=True)[:top_n]
            x_cats = [k for k, v in sorted_items]
            y_vals = [v for k, v in sorted_items]
            title = f"Top {top_n} {dim_col.name} by {measure_col.name}"
        else:
            x_cats = list(summary_dict.keys())
            y_vals = list(summary_dict.values())
            title = f"{measure_col.name} by {dim_col.name}"

        chart_id = f"chart_{uuid.uuid4().hex[:8]}"
        dataset_dir = file_manager.get_dataset_dir(workbook_index.dataset_id)
        output_file = dataset_dir / "charts" / f"{chart_id}.png"

        series_spec = ChartSeriesSpec(name=f"{agg_name} {measure_col.name}", values=y_vals)
        base64_str = ChartRenderer.render_to_file(
            chart_type=chart_type,
            title=title,
            x_categories=x_cats,
            series=[series_spec],
            x_axis_label=dim_col.name,
            y_axis_label=measure_col.name,
            output_file_path=output_file,
            include_base64=True,
        )
        image_url = f"/api/v1/datasets/{workbook_index.dataset_id}/charts/{chart_id}/image"

        return ChartActionSpec(
            chart_id=chart_id,
            chart_type=chart_type.value,
            title=title,
            dimension_column=dim_col.name,
            measure_column=measure_col.name,
            aggregation=agg_name,
            destination_cell=dest_cell,
            width_cols=width_cols,
            height_rows=height_rows,
            image_url=image_url,
            image_base64=base64_str,
            summary_data=[{"category": c, "value": v} for c, v in zip(x_cats, y_vals)],
            source_range=table_entry.range_address,
            provenance_note=f"Aggregated {len(grouped)} categories from table '{table_entry.name}' via Python {agg_name} on {measure_col.name}",
        )

    @classmethod
    def _resolve_default_chart_anchor(cls, table_entry: TableIndexEntry, grid: Optional[RawSheetGrid] = None) -> str:
        """Deterministically calculates fallback chart anchor below table data and existing charts."""
        max_r = table_entry.row_count + 1
        if table_entry.data_range and ":" in table_entry.data_range:
            try:
                _, end_c = table_entry.data_range.split(":")
                _, max_r = coordinate_from_string(end_c)
            except Exception:
                pass
        start_row = max(max_r + 2, 8)

        if grid and grid.charts:
            max_chart_bottom = 0
            for ex_c in grid.charts.values():
                dest = ex_c.get("destination_cell") or ex_c.get("anchor_cell")
                if dest:
                    try:
                        _, r_int = coordinate_from_string(dest.strip().upper())
                        h = ex_c.get("height_rows", 14)
                        max_chart_bottom = max(max_chart_bottom, r_int + h)
                    except Exception:
                        pass
            if max_chart_bottom > 0:
                start_row = max(start_row, max_chart_bottom + 2)

        return f"B{start_row}"

    @classmethod
    def _plan_multi_visualizations(
        cls,
        grid: RawSheetGrid,
        workbook_index: WorkbookMetadataIndex,
        cur_sheet: str,
        table_entry: TableIndexEntry,
        q_norm: str,
        explicit_dest_cell: Optional[str] = None,
        selected_range: Optional[str] = None,
        target_range_context: Optional[str] = None,
    ) -> Tuple[List[SpreadsheetAction], AgentResponseStatusEnum, Optional[ClarificationRequest], str]:
        """Deterministically generates a bounded set of relevant, non-redundant visualizations in a clean grid layout."""
        # 1. Resolve starting anchor
        start_dest = explicit_dest_cell or (selected_range.split(":")[0].upper() if selected_range else None) or cls._resolve_default_chart_anchor(table_entry)
        try:
            start_col_letter, start_row = coordinate_from_string(start_dest)
            start_col_idx = column_index_from_string(start_col_letter)
        except Exception:
            start_col_idx = 2
            start_row = 12

        # 2. Extract column candidates
        num_cols = [
            c for c in table_entry.columns
            if (c.semantic_type == SemanticTypeEnum.NUMERIC_MEASURE or c.data_type in {DataTypeEnum.FLOAT, DataTypeEnum.CURRENCY, DataTypeEnum.PERCENTAGE, DataTypeEnum.INTEGER})
            and c.semantic_type != SemanticTypeEnum.IDENTIFIER
            and not c.name.lower().endswith("id")
            and not c.name.lower().startswith("id")
        ]
        if not num_cols:
            num_cols = [c for c in table_entry.columns if c.data_type in {DataTypeEnum.INTEGER, DataTypeEnum.FLOAT}]
        if not num_cols:
            num_cols = [table_entry.columns[-1]]

        cat_cols = [
            c for c in table_entry.columns
            if (c.semantic_type in {SemanticTypeEnum.CATEGORICAL, SemanticTypeEnum.IDENTIFIER} or c.data_type == DataTypeEnum.STRING)
            and not c.name.lower().endswith("id")
            and c not in num_cols
        ]
        if not cat_cols:
            cat_cols = [c for c in table_entry.columns if c.data_type == DataTypeEnum.STRING and c not in num_cols]
        if not cat_cols:
            cat_cols = [table_entry.columns[0]]

        temp_cols = [
            c for c in table_entry.columns
            if c.semantic_type == SemanticTypeEnum.TEMPORAL or c.data_type in {DataTypeEnum.DATE, DataTypeEnum.DATETIME}
            or any(kw in c.normalized_name for kw in ["date", "tanggal", "bulan", "tahun", "year", "month", "time"])
        ]

        # 3. Plan bounded chart candidates (3 to 6 charts)
        chart_plans: List[Tuple[ColumnIndexEntry, ColumnIndexEntry, ChartTypeEnum, str]] = []

        m1 = num_cols[0]
        dim1 = cat_cols[0]
        c1_type = ChartTypeEnum.PIE if len(cat_cols) > 0 else ChartTypeEnum.BAR
        chart_plans.append((dim1, m1, c1_type, "SUM"))

        if len(cat_cols) > 1:
            dim2 = cat_cols[1]
            chart_plans.append((dim2, m1, ChartTypeEnum.COLUMN, "SUM"))
        elif len(num_cols) > 1:
            m2 = num_cols[1]
            chart_plans.append((dim1, m2, ChartTypeEnum.BAR, "SUM"))

        if temp_cols:
            t_dim = temp_cols[0]
            chart_plans.append((t_dim, m1, ChartTypeEnum.LINE, "SUM"))
        elif len(num_cols) > 1 and len(cat_cols) > 1:
            m2 = num_cols[1]
            chart_plans.append((dim1, m2, ChartTypeEnum.BAR, "SUM"))

        if len(num_cols) > 1:
            m2 = num_cols[1]
            if len(cat_cols) > 1 and (cat_cols[1], m2, ChartTypeEnum.COLUMN, "SUM") not in chart_plans:
                chart_plans.append((cat_cols[1], m2, ChartTypeEnum.COLUMN, "SUM"))
            elif temp_cols and (temp_cols[0], m2, ChartTypeEnum.LINE, "SUM") not in chart_plans:
                chart_plans.append((temp_cols[0], m2, ChartTypeEnum.LINE, "SUM"))

        seen_keys = set()
        unique_plans = []
        for d_col, m_col, ct, agg in chart_plans:
            k = (d_col.name.strip().lower(), m_col.name.strip().lower(), ct.value.strip().lower())
            if k not in seen_keys:
                seen_keys.add(k)
                # Deduplication check against existing charts on the worksheet
                is_already_present = False
                if grid and grid.charts:
                    for ex_id, ex_c in grid.charts.items():
                        ex_dim = (ex_c.get("dimension_column") or "").strip().lower()
                        ex_meas = (ex_c.get("measure_column") or "").strip().lower()
                        ex_type = (ex_c.get("chart_type") or "").strip().lower()
                        if ex_dim == d_col.name.strip().lower() and ex_meas == m_col.name.strip().lower() and ex_type == ct.value.strip().lower():
                            is_already_present = True
                            break
                if not is_already_present:
                    unique_plans.append((d_col, m_col, ct, agg))

        is_english = any(w in q_norm.split() for w in ["all", "visualize", "everything", "useful", "show"])

        # If all candidates already exist, return 0 mutations with informative message
        if not unique_plans:
            msg = (
                "All relevant visualizations already exist on the worksheet, so no duplicate charts were created."
                if is_english
                else "Semua visualisasi yang relevan sudah tersedia di lembar kerja, jadi saya tidak membuat duplikat."
            )
            return [], AgentResponseStatusEnum.SUCCESS, None, msg

        # 4. Generate actions in a clean, non-overlapping 2-column grid layout
        actions: List[SpreadsheetAction] = []
        width_cols = 7
        height_rows = 14
        col_gap = 8
        row_gap = 16

        # Track occupied boxes (both existing charts and newly placed charts)
        occupied_boxes: List[Tuple[int, int, int, int]] = []
        if grid and grid.charts:
            for ex_id, ex_c in grid.charts.items():
                ex_dest = ex_c.get("destination_cell") or ex_c.get("anchor_cell")
                if ex_dest:
                    try:
                        ex_col_str, ex_row = coordinate_from_string(ex_dest.strip().upper())
                        ex_col = column_index_from_string(ex_col_str)
                        ex_w = ex_c.get("width_cols", 7)
                        ex_h = ex_c.get("height_rows", 14)
                        occupied_boxes.append((ex_col, ex_row, ex_col + ex_w - 1, ex_row + ex_h - 1))
                    except Exception:
                        pass

        def is_box_colliding(c_box: Tuple[int, int, int, int]) -> bool:
            for ex_box in occupied_boxes:
                if not (c_box[2] < ex_box[0] or c_box[0] > ex_box[2] or c_box[3] < ex_box[1] or c_box[1] > ex_box[3]):
                    return True
            return False

        for idx, (d_col, m_col, ct, agg) in enumerate(unique_plans):
            # Find a free non-overlapping position
            slot_found = False
            slot_idx = idx
            dest_coord = ""
            while not slot_found and slot_idx < 100:
                col_offset = (slot_idx % 2) * col_gap
                row_offset = (slot_idx // 2) * row_gap
                c_col_int = start_col_idx + col_offset
                c_row_int = start_row + row_offset
                cand_box = (c_col_int, c_row_int, c_col_int + width_cols - 1, c_row_int + height_rows - 1)
                if not is_box_colliding(cand_box):
                    slot_found = True
                    occupied_boxes.append(cand_box)
                    dest_coord = f"{get_column_letter(c_col_int)}{c_row_int}"
                else:
                    slot_idx += 1

            if not dest_coord:
                dest_coord = f"{get_column_letter(start_col_idx)}{start_row + idx * row_gap}"

            chart_spec = cls._build_chart_spec(
                grid=grid,
                workbook_index=workbook_index,
                cur_sheet=cur_sheet,
                table_entry=table_entry,
                dim_col=d_col,
                measure_col=m_col,
                chart_type=ct,
                agg_name=agg,
                dest_cell=dest_coord,
                target_range_context=target_range_context,
                width_cols=width_cols,
                height_rows=height_rows,
            )

            actions.append(SpreadsheetAction(
                action_id=f"act_{idx + 1}",
                action_type=ActionTypeEnum.CREATE_CHART,
                sheet_name=cur_sheet,
                target_cell=dest_coord,
                chart_spec=chart_spec,
                description=f"Create {ct.value} chart '{chart_spec.title}' at {dest_coord}",
            ))

        val_res = ActionValidator.validate_sequence(actions, workbook_index, grid)
        if not val_res.is_valid:
            return [], val_res.status, val_res.clarification_request, val_res.error_message or "Validation failed."

        msg = (
            f"Done. Created {len(actions)} relevant visualizations based on the data structure."
            if is_english
            else f"Selesai. Saya membuat {len(actions)} visualisasi yang relevan berdasarkan struktur data."
        )
        return actions, AgentResponseStatusEnum.SUCCESS, None, msg

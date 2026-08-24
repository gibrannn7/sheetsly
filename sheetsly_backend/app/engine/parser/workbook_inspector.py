"""Workbook inspector for reading multi-sheet workbooks and compiling RawSheetGrid representations."""

from pathlib import Path
from typing import Dict, List, Tuple
import openpyxl

from app.core.errors import WorkbookParseError
from app.core.logging import logger
from .sheet_reader import RawSheetGrid, SheetReader


class WorkbookInspector:
    """Deterministic parser that inspects complete workbook files without calling external AI."""

    @classmethod
    def inspect_file(cls, file_path: Path) -> Tuple[List[str], Dict[str, RawSheetGrid]]:
        """
        Inspects the workbook file and loads each sheet into a RawSheetGrid.
        Returns: (sheet_names, dict_of_grids)
        """
        if not file_path.exists():
            raise WorkbookParseError(f"Workbook file does not exist: {file_path}")

        ext = file_path.suffix.lower()
        if ext == ".csv":
            sheet_name = "Sheet1"
            grid = SheetReader.read_csv_file(file_path, sheet_name=sheet_name)
            return [sheet_name], {sheet_name: grid}

        try:
            # Load evaluated workbook (values after formulas calculated in Excel)
            wb_eval = openpyxl.load_workbook(
                filename=str(file_path),
                data_only=True,
                read_only=False,
                keep_vba=True if ext == ".xlsm" else False,
            )

            # Load formula workbook (contains raw formula strings)
            try:
                wb_formula = openpyxl.load_workbook(
                    filename=str(file_path),
                    data_only=False,
                    read_only=False,
                    keep_vba=True if ext == ".xlsm" else False,
                )
            except Exception as e:
                logger.warning(f"Could not load formula view for {file_path.name}: {e}")
                wb_formula = None

            sheet_names = wb_eval.sheetnames
            if not sheet_names:
                raise WorkbookParseError("Workbook contains no sheets.")

            grids: Dict[str, RawSheetGrid] = {}
            for name in sheet_names:
                sheet_eval = wb_eval[name]
                sheet_formula = wb_formula[name] if wb_formula and name in wb_formula.sheetnames else None
                grid = SheetReader.read_openpyxl_worksheet(sheet_eval, sheet_formula)
                grids[name] = grid

            # Close workbooks
            try:
                wb_eval.close()
                if wb_formula:
                    wb_formula.close()
            except Exception:
                pass

            logger.info(f"Inspected workbook '{file_path.name}' with {len(sheet_names)} sheets: {sheet_names}")
            return sheet_names, grids

        except Exception as e:
            if isinstance(e, WorkbookParseError):
                raise
            raise WorkbookParseError(f"Failed to inspect Excel workbook '{file_path.name}': {str(e)}")

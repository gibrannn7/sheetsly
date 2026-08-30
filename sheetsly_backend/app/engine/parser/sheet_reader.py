"""Sheet reader preserving raw cell coordinates, values, formulas, and merged regions."""

import csv
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

from app.core.errors import WorkbookParseError
from app.models.schemas import CellCoordinate, CellData, DataTypeEnum


@dataclass
class RawSheetGrid:
    """In-memory deterministic 2D representation of a sheet with cell coordinates."""

    sheet_name: str
    min_row: int = 1
    max_row: int = 1
    min_col: int = 1
    max_col: int = 1
    cells: Dict[Tuple[int, int], CellData] = field(default_factory=dict)
    merged_ranges: List[str] = field(default_factory=list)
    charts: Dict[str, Any] = field(default_factory=dict)
    kpis: Dict[str, Any] = field(default_factory=dict)
    is_hidden: bool = False

    @property
    def total_rows(self) -> int:
        return max(0, self.max_row - self.min_row + 1) if self.max_row >= self.min_row else 0

    @property
    def total_cols(self) -> int:
        return max(0, self.max_col - self.min_col + 1) if self.max_col >= self.min_col else 0

    @property
    def used_range(self) -> str:
        if self.total_rows == 0 or self.total_cols == 0:
            return "A1:A1"
        start_col = get_column_letter(self.min_col)
        end_col = get_column_letter(self.max_col)
        return f"{start_col}{self.min_row}:{end_col}{self.max_row}"

    def get_cell(self, row: int, col: int) -> CellData:
        """Returns CellData at 1-indexed (row, col), returning an empty CellData if not present."""
        if (row, col) in self.cells:
            return self.cells[(row, col)]
        col_letter = get_column_letter(col)
        return CellData(
            coordinate=CellCoordinate(row=row, column=col, cell_ref=f"{col_letter}{row}"),
            original_value=None,
            parsed_value=None,
            data_type=DataTypeEnum.NULL,
            formula=None,
            is_empty=True,
        )

    def get_row_cells(self, row: int, min_col: Optional[int] = None, max_col: Optional[int] = None) -> List[CellData]:
        """Returns ordered list of cells for a given row number."""
        c_start = min_col or self.min_col
        c_end = max_col or self.max_col
        return [self.get_cell(row, c) for c in range(c_start, c_end + 1)]

    def get_col_cells(self, col: int, min_row: Optional[int] = None, max_row: Optional[int] = None) -> List[CellData]:
        """Returns ordered list of cells for a given column number."""
        r_start = min_row or self.min_row
        r_end = max_row or self.max_row
        return [self.get_cell(r, col) for r in range(r_start, r_end + 1)]

    def is_row_empty(self, row: int, min_col: Optional[int] = None, max_col: Optional[int] = None) -> bool:
        """True if all cells in the row segment are empty/null."""
        cells = self.get_row_cells(row, min_col, max_col)
        return all(c.is_empty for c in cells)

    def is_col_empty(self, col: int, min_row: Optional[int] = None, max_row: Optional[int] = None) -> bool:
        """True if all cells in the column segment are empty/null."""
        cells = self.get_col_cells(col, min_row, max_row)
        return all(c.is_empty for c in cells)

    def get_2d_slice(
        self, min_row: int, min_col: int, max_row: int, max_col: int
    ) -> List[List[CellData]]:
        """Returns a 2D matrix of CellData within the specified bounding box."""
        matrix = []
        for r in range(min_row, max_row + 1):
            row_cells = [self.get_cell(r, c) for c in range(min_col, max_col + 1)]
            matrix.append(row_cells)
        return matrix


class SheetReader:
    """Reads Excel and CSV worksheets into deterministic RawSheetGrid structures."""

    @staticmethod
    def read_openpyxl_worksheet(
        sheet_evaluated: Worksheet,
        sheet_formula: Optional[Worksheet] = None,
    ) -> RawSheetGrid:
        """
        Reads openpyxl worksheet into a RawSheetGrid.
        Takes evaluated sheet (data_only=True) and formula sheet (data_only=False).
        """
        sheet_name = sheet_evaluated.title
        is_hidden = getattr(sheet_evaluated, "sheet_state", "visible") == "hidden"
        merged_ranges = [str(r) for r in sheet_evaluated.merged_cells.ranges]

        # Determine non-empty used bounds
        min_r, max_r = 1, 1
        min_c, max_c = 1, 1
        has_any_data = False

        # Gather cell coordinates and values
        cell_dict: Dict[Tuple[int, int], CellData] = {}

        # Scan all populated cells in sheet
        for row in sheet_evaluated.iter_rows(values_only=False):
            for cell in row:
                r = cell.row
                c = cell.column
                val_eval = cell.value

                formula_str = None
                if sheet_formula is not None:
                    try:
                        raw_cell = sheet_formula.cell(row=r, column=c)
                        if raw_cell.value and isinstance(raw_cell.value, str) and raw_cell.value.startswith("="):
                            formula_str = str(raw_cell.value)
                    except Exception:
                        pass

                is_empty = val_eval is None or (isinstance(val_eval, str) and val_eval.strip() == "")
                if not is_empty or formula_str is not None:
                    if not has_any_data:
                        min_r, max_r = r, r
                        min_c, max_c = c, c
                        has_any_data = True
                    else:
                        min_r = min(min_r, r)
                        max_r = max(max_r, r)
                        min_c = min(min_c, c)
                        max_c = max(max_c, c)

                col_letter = get_column_letter(c)
                coord = CellCoordinate(row=r, column=c, cell_ref=f"{col_letter}{r}")

                cell_dict[(r, c)] = CellData(
                    coordinate=coord,
                    original_value=val_eval,
                    parsed_value=val_eval,
                    data_type=DataTypeEnum.UNKNOWN,
                    formula=formula_str,
                    is_empty=is_empty,
                )

        if not has_any_data:
            # Empty sheet
            return RawSheetGrid(
                sheet_name=sheet_name,
                min_row=1,
                max_row=1,
                min_col=1,
                max_col=1,
                cells={},
                merged_ranges=merged_ranges,
                is_hidden=is_hidden,
            )

        return RawSheetGrid(
            sheet_name=sheet_name,
            min_row=min_r,
            max_row=max_r,
            min_col=min_c,
            max_col=max_c,
            cells=cell_dict,
            merged_ranges=merged_ranges,
            is_hidden=is_hidden,
        )

    @staticmethod
    def read_csv_file(file_path: Path, sheet_name: str = "Sheet1") -> RawSheetGrid:
        """Reads a CSV file into a RawSheetGrid."""
        cell_dict: Dict[Tuple[int, int], CellData] = {}
        min_r, max_r = 1, 1
        min_c, max_c = 1, 1
        has_any_data = False

        try:
            # Try utf-8 first, fallback to latin-1
            encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]
            rows_data = None
            for enc in encodings:
                try:
                    with open(file_path, "r", encoding=enc, newline="") as f:
                        # Sniff delimiter
                        sample = f.read(4096)
                        f.seek(0)
                        try:
                            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
                            delimiter = dialect.delimiter
                        except Exception:
                            delimiter = ","
                        reader = csv.reader(f, delimiter=delimiter)
                        rows_data = list(reader)
                        break
                except UnicodeDecodeError:
                    continue

            if rows_data is None:
                raise WorkbookParseError(f"Could not decode CSV file '{file_path.name}'.")

            for r_idx, row in enumerate(rows_data, start=1):
                for c_idx, val in enumerate(row, start=1):
                    is_empty = val is None or str(val).strip() == ""
                    if not is_empty:
                        if not has_any_data:
                            min_r, max_r = r_idx, r_idx
                            min_c, max_c = c_idx, c_idx
                            has_any_data = True
                        else:
                            min_r = min(min_r, r_idx)
                            max_r = max(max_r, r_idx)
                            min_c = min(min_c, c_idx)
                            max_c = max(max_c, c_idx)

                    col_letter = get_column_letter(c_idx)
                    coord = CellCoordinate(row=r_idx, column=c_idx, cell_ref=f"{col_letter}{r_idx}")
                    cell_dict[(r_idx, c_idx)] = CellData(
                        coordinate=coord,
                        original_value=val if not is_empty else None,
                        parsed_value=val if not is_empty else None,
                        data_type=DataTypeEnum.STRING if not is_empty else DataTypeEnum.NULL,
                        formula=None,
                        is_empty=is_empty,
                    )

            return RawSheetGrid(
                sheet_name=sheet_name,
                min_row=min_r,
                max_row=max_r,
                min_col=min_c,
                max_col=max_c,
                cells=cell_dict,
                merged_ranges=[],
                is_hidden=False,
            )

        except Exception as e:
            if isinstance(e, WorkbookParseError):
                raise
            raise WorkbookParseError(f"Failed to read CSV spreadsheet: {str(e)}")

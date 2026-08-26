import csv
import io
from io import BytesIO
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi import APIRouter, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse

from app.core.errors import FileValidationError
from app.core.logging import logger
from app.engine.pipeline import ingestion_pipeline
from app.models.schemas import WorkbookOverview
from app.storage.file_manager import file_manager

router = APIRouter(prefix="/datasets", tags=["Datasets"])


@router.post("/upload", response_model=WorkbookOverview, status_code=status.HTTP_201_CREATED)
async def upload_spreadsheet(
    file: UploadFile = File(..., description="Spreadsheet file (.xlsx, .xls, .csv, .xlsm)"),
) -> WorkbookOverview:
    """
    Ingests and deterministically inspects an uploaded spreadsheet workbook.
    Returns complete structural metadata, sheet inventory, detected tables, column types, and data quality metrics.
    """
    if not file.filename:
        raise FileValidationError("Filename cannot be empty.")

    # 1. Save uploaded file to secure temporary storage
    dataset_id, file_path, original_filename, file_size_bytes = await file_manager.save_uploaded_file(file)

    # 2. Run deterministic inspection pipeline
    try:
        overview = ingestion_pipeline.process_workbook(
            dataset_id=dataset_id,
            file_path=file_path,
            original_filename=original_filename,
            file_size_bytes=file_size_bytes,
        )
        return overview
    except Exception as e:
        # Clean up stored file if processing failed
        file_manager.cleanup_dataset(dataset_id)
        raise


@router.get("/{dataset_id}", response_model=WorkbookOverview)
async def get_dataset_overview(dataset_id: str) -> WorkbookOverview:
    """Retrieves workbook structural overview and inspection report for an active dataset ID."""
    return ingestion_pipeline.get_overview(dataset_id)


@router.get("/{dataset_id}/export")
async def export_dataset(
    dataset_id: str,
    format: str = Query("xlsx", pattern="^(xlsx|csv)$", description="Export format: 'xlsx' (full workbook) or 'csv' (active sheet)"),
    sheet_name: Optional[str] = Query(None, description="Optional target sheet name (primarily for CSV export)"),
):
    """
    Exports the live, mutated workbook representation as a full-fidelity Excel (.xlsx) file
    or active-sheet CSV (.csv) file, preserving all worksheets, formula strings, cell values, and mutations.
    """
    overview = ingestion_pipeline.get_overview(dataset_id)
    grids = ingestion_pipeline.get_all_grids(dataset_id)

    if not grids:
        raise HTTPException(status_code=404, detail="No active grid found for dataset.")

    base_name = overview.filename.rsplit(".", 1)[0] if "." in overview.filename else overview.filename

    if format == "xlsx":
        wb = openpyxl.Workbook()
        default_ws = wb.active

        for s_idx, (s_name, grid) in enumerate(grids.items()):
            ws = wb.create_sheet(title=s_name)
            for r in range(grid.min_row, grid.max_row + 1):
                for c in range(grid.min_col, grid.max_col + 1):
                    cell_data = grid.get_cell(r, c)
                    if cell_data and not cell_data.is_empty:
                        dest_cell = ws.cell(row=r, column=c)
                        if cell_data.formula:
                            dest_cell.value = cell_data.formula
                        elif cell_data.parsed_value is not None:
                            dest_cell.value = cell_data.parsed_value
                        else:
                            dest_cell.value = cell_data.original_value

                        style = getattr(cell_data, "style", None)
                        if style:
                            font_kwargs = {}
                            if getattr(style, "bold", None) is not None:
                                font_kwargs["bold"] = style.bold
                            if getattr(style, "italic", None) is not None:
                                font_kwargs["italic"] = style.italic
                            if getattr(style, "font_size", None):
                                font_kwargs["size"] = style.font_size
                            if getattr(style, "font_color", None):
                                font_kwargs["color"] = style.font_color.replace("#", "")
                            if font_kwargs:
                                dest_cell.font = Font(**font_kwargs)

                            fill_color = getattr(style, "fill_color", None)
                            if fill_color:
                                fill_hex = fill_color.replace("#", "")
                                dest_cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")

                        num_fmt = getattr(cell_data, "number_format", None)
                        if num_fmt:
                            dest_cell.number_format = num_fmt

        if default_ws in wb.worksheets and len(wb.worksheets) > 1:
            wb.remove(default_ws)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        export_filename = f"{base_name}.xlsx"

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{export_filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition",
            },
        )

    else: # CSV format
        target_sheet = sheet_name or (overview.sheets[0].name if overview.sheets else list(grids.keys())[0])
        if target_sheet not in grids:
            raise HTTPException(status_code=404, detail=f"Sheet '{target_sheet}' not found in dataset.")

        grid = grids[target_sheet]
        output = io.StringIO()
        writer = csv.writer(output)

        for r in range(grid.min_row, grid.max_row + 1):
            row_vals = []
            for c in range(grid.min_col, grid.max_col + 1):
                cell_data = grid.get_cell(r, c)
                if cell_data and not cell_data.is_empty:
                    if cell_data.formula:
                        row_vals.append(cell_data.formula)
                    elif cell_data.parsed_value is not None:
                        row_vals.append(cell_data.parsed_value)
                    else:
                        row_vals.append(cell_data.original_value)
                else:
                    row_vals.append("")
            writer.writerow(row_vals)

        csv_bytes = output.getvalue().encode("utf-8")
        export_filename = f"{base_name}_{target_sheet}.csv" if len(grids) > 1 else f"{base_name}.csv"

        return StreamingResponse(
            BytesIO(csv_bytes),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{export_filename}"',
                "Access-Control-Expose-Headers": "Content-Disposition",
            },
        )


@router.delete("/{dataset_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_dataset(dataset_id: str) -> None:
    """Cleans up temporary files and session data for a dataset ID."""
    file_manager.cleanup_dataset(dataset_id)

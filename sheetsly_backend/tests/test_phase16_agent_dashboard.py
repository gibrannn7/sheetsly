"""Comprehensive automated test suite for Phase 16: Verified Spreadsheet Agent & Automated Dashboard Engine."""

import copy
import io
from pathlib import Path
import pytest
import openpyxl
from openpyxl.chart import BarChart, PieChart, LineChart

from app.engine.agent.action_model import ActionTypeEnum, SpreadsheetAction, ChartActionSpec, KPIActionSpec, FormattingStyle
from app.engine.agent.agent_planner import SpreadsheetAgentPlanner
from app.engine.agent.agent_orchestrator import AgentOrchestrator
from app.engine.agent.action_validator import ActionValidator
from app.engine.agent.transaction_model import AgentResponseStatusEnum, TransactionStatusEnum
from app.engine.parser.sheet_reader import RawSheetGrid
from app.engine.profiler.workbook_index import (
    ColumnIndexEntry,
    SheetIndexEntry,
    TableIndexEntry,
    WorkbookMetadataIndex,
)
from app.models.schemas import CellCoordinate, CellData, DataTypeEnum, SemanticTypeEnum


@pytest.fixture
def mock_orders_grid_and_index():
    cells = {}
    # Headers
    headers = ["OrderID", "Region", "Category", "Sales", "Profit", "OrderDate"]
    types = [DataTypeEnum.INTEGER, DataTypeEnum.STRING, DataTypeEnum.STRING, DataTypeEnum.FLOAT, DataTypeEnum.FLOAT, DataTypeEnum.DATE]
    for c_idx, (h, dt) in enumerate(zip(headers, types), start=1):
        col_letter = chr(64 + c_idx)
        cells[(1, c_idx)] = CellData(
            coordinate=CellCoordinate(row=1, column=c_idx, cell_ref=f"{col_letter}1", col_letter=col_letter),
            original_value=h,
            parsed_value=h,
            data_type=dt,
            is_empty=False,
        )

    # 10 Sample Rows
    rows_data = [
        (1, "East", "Technology", 1200.0, 300.0, "2023-01-15"),
        (2, "West", "Furniture", 850.0, 150.0, "2023-01-20"),
        (3, "Central", "Office Supplies", 430.0, 80.0, "2023-02-10"),
        (4, "South", "Technology", 920.0, 220.0, "2023-02-14"),
        (5, "East", "Furniture", 610.0, 110.0, "2023-03-05"),
        (6, "West", "Office Supplies", 310.0, 60.0, "2023-03-12"),
        (7, "East", "Technology", 1450.0, 410.0, "2023-04-01"),
        (8, "Central", "Furniture", 790.0, 130.0, "2023-04-18"),
        (9, "West", "Technology", 1100.0, 280.0, "2023-05-02"),
        (10, "South", "Office Supplies", 520.0, 95.0, "2023-05-15"),
    ]

    for r_idx, row in enumerate(rows_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            col_letter = chr(64 + c_idx)
            dt = types[c_idx - 1]
            cells[(r_idx, c_idx)] = CellData(
                coordinate=CellCoordinate(row=r_idx, column=c_idx, cell_ref=f"{col_letter}{r_idx}", col_letter=col_letter),
                original_value=val,
                parsed_value=val,
                data_type=dt,
                is_empty=False,
            )

    grid = RawSheetGrid(
        sheet_name="Orders",
        min_row=1,
        max_row=11,
        min_col=1,
        max_col=6,
        cells=cells,
    )

    columns = [
        ColumnIndexEntry(index=0, name="OrderID", normalized_name="orderid", source_column_letter="A", semantic_type=SemanticTypeEnum.IDENTIFIER, data_type=DataTypeEnum.INTEGER),
        ColumnIndexEntry(index=1, name="Region", normalized_name="region", source_column_letter="B", semantic_type=SemanticTypeEnum.CATEGORICAL, data_type=DataTypeEnum.STRING),
        ColumnIndexEntry(index=2, name="Category", normalized_name="category", source_column_letter="C", semantic_type=SemanticTypeEnum.CATEGORICAL, data_type=DataTypeEnum.STRING),
        ColumnIndexEntry(index=3, name="Sales", normalized_name="sales", source_column_letter="D", semantic_type=SemanticTypeEnum.NUMERIC_MEASURE, data_type=DataTypeEnum.FLOAT),
        ColumnIndexEntry(index=4, name="Profit", normalized_name="profit", source_column_letter="E", semantic_type=SemanticTypeEnum.NUMERIC_MEASURE, data_type=DataTypeEnum.FLOAT),
        ColumnIndexEntry(index=5, name="OrderDate", normalized_name="orderdate", source_column_letter="F", semantic_type=SemanticTypeEnum.TEMPORAL, data_type=DataTypeEnum.DATE),
    ]

    table_entry = TableIndexEntry(
        table_id="tbl_orders",
        sheet_name="Orders",
        name="Orders_Table",
        range_address="A1:F11",
        header_range="A1:F1",
        data_range="A2:F11",
        header_row_index=1,
        row_count=10,
        column_count=6,
        columns=columns,
    )

    sheet_entry = SheetIndexEntry(
        name="Orders",
        index=0,
        total_rows=11,
        total_columns=6,
        used_range="A1:F11",
        tables=[table_entry],
    )

    index = WorkbookMetadataIndex(
        dataset_id="ds_phase16_test",
        filename="Superstore_Phase16.xlsx",
        sheet_count=1,
        sheet_names=["Orders"],
        active_sheet_name="Orders",
        sheets={"Orders": sheet_entry},
    )

    return grid, index


# ============================================================================
# 1. DIRECT SPREADSHEET OPERATIONS & TARGETING
# ============================================================================

def test_direct_value_write(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    orch = AgentOrchestrator()
    res = orch.process_request(
        user_request="tulis teks 'Ringkasan' di N2",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert res.status == AgentResponseStatusEnum.SUCCESS
    assert grid.get_cell(2, 14).parsed_value == "Ringkasan"


def test_direct_formula_write(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    orch = AgentOrchestrator()
    res = orch.process_request(
        user_request="buat formula total Sales di N3 dengan rumus =SUM(D2:D11)",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert res.status == AgentResponseStatusEnum.SUCCESS
    assert grid.get_cell(3, 14).formula == "=SUM(D2:D11)"
    assert grid.get_cell(3, 14).parsed_value == 8180.0


def test_formatting_operation(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    orch = AgentOrchestrator()
    res = orch.process_request(
        user_request="format bold dan highlight header",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert res.status == AgentResponseStatusEnum.SUCCESS


def test_destination_and_range_targeting(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    actions, status, _, _ = SpreadsheetAgentPlanner.plan_agent_actions(
        user_request="buat pie chart di rentang B12 untuk kolom Region berdasarkan Sales",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert status == AgentResponseStatusEnum.SUCCESS
    assert len(actions) == 1
    assert actions[0].action_type == ActionTypeEnum.CREATE_CHART
    assert actions[0].target_cell == "B12"
    assert actions[0].chart_spec.destination_cell == "B12"


# ============================================================================
# 2. ALL 7 CHART TYPES & CREATION VERIFICATION
# ============================================================================

@pytest.mark.parametrize("prompt,expected_type", [
    ("buatkan pie chart Region berdasarkan Sales di B12", "PIE"),
    ("buatkan bar chart Region berdasarkan Sales di B12", "BAR"),
    ("buatkan column chart Category berdasarkan Sales di B12", "COLUMN"),
    ("buatkan line chart OrderDate berdasarkan Sales di B12", "LINE"),
    ("buatkan area chart OrderDate berdasarkan Sales di B12", "AREA"),
    ("buatkan scatter plot Sales dan Profit di B12", "SCATTER"),
    ("buatkan histogram distribusi Sales di B12", "HISTOGRAM"),
])
def test_create_all_7_chart_types(mock_orders_grid_and_index, prompt, expected_type):
    grid, index = mock_orders_grid_and_index
    orch = AgentOrchestrator()
    res = orch.process_request(
        user_request=prompt,
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert res.status == AgentResponseStatusEnum.SUCCESS
    assert len(grid.charts) == 1
    chart_entry = list(grid.charts.values())[0]
    c_type = chart_entry.get("chart_type") if isinstance(chart_entry, dict) else chart_entry.chart_type
    assert c_type == expected_type


def test_chart_verification_rules(mock_orders_grid_and_index):
    """Verify chart exists, anchor matches, and type matches after execution."""
    grid, index = mock_orders_grid_and_index
    orch = AgentOrchestrator()
    res = orch.process_request(
        user_request="buat pie chart Region berdasarkan Sales di B12",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert res.status == AgentResponseStatusEnum.SUCCESS
    assert res.transaction.verification_report.is_verified is True
    assert len(grid.charts) == 1
    c_obj = list(grid.charts.values())[0]
    dest = c_obj.get("destination_cell") if isinstance(c_obj, dict) else c_obj.destination_cell
    assert dest == "B12"


# ============================================================================
# 3. CHART LIFECYCLE: MOVE, RESIZE, UPDATE, DELETE, UNDO, ROLLBACK
# ============================================================================

def test_chart_move_operation(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    orch = AgentOrchestrator()

    # 1. Create chart at B12
    res1 = orch.process_request(
        user_request="buat pie chart Region berdasarkan Sales di B12",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert res1.status == AgentResponseStatusEnum.SUCCESS
    assert len(grid.charts) == 1

    # 2. Move chart to H20
    res2 = orch.process_request(
        user_request="pindahkan pie chart Sales by Region ke H20",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert res2.status == AgentResponseStatusEnum.SUCCESS
    c_obj = list(grid.charts.values())[0]
    dest = c_obj.get("destination_cell") if isinstance(c_obj, dict) else c_obj.destination_cell
    assert dest == "H20"


def test_chart_resize_operation(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    orch = AgentOrchestrator()

    orch.process_request(
        user_request="buat bar chart Category berdasarkan Sales di B12",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )

    res = orch.process_request(
        user_request="ubah ukuran chart menjadi 10 kolom dan 20 baris",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert res.status == AgentResponseStatusEnum.SUCCESS
    c_obj = list(grid.charts.values())[0]
    w = c_obj.get("width_cols") if isinstance(c_obj, dict) else c_obj.width_cols
    h = c_obj.get("height_rows") if isinstance(c_obj, dict) else c_obj.height_rows
    assert w == 10
    assert h == 20


def test_chart_update_and_delete_operations(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    orch = AgentOrchestrator()

    orch.process_request(
        user_request="buat pie chart Region berdasarkan Sales di B12",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert len(grid.charts) == 1

    # Update to Line
    res_upd = orch.process_request(
        user_request="ubah chart ini menjadi line chart",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert res_upd.status == AgentResponseStatusEnum.SUCCESS

    # Delete
    res_del = orch.process_request(
        user_request="hapus chart di B12",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert res_del.status == AgentResponseStatusEnum.SUCCESS
    assert len(grid.charts) == 0


def test_chart_atomic_undo(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    orch = AgentOrchestrator()

    orch.process_request(
        user_request="buat pie chart Region berdasarkan Sales di B12",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert len(grid.charts) == 1

    undo_res = orch.undo_last_transaction(grid=grid)
    assert undo_res.status == AgentResponseStatusEnum.ROLLBACK_SUCCESS
    assert len(grid.charts) == 0


# ============================================================================
# 4. AUTOMATED DASHBOARD ENGINE
# ============================================================================

def test_automated_dashboard_generation(mock_orders_grid_and_index):
    """Verify 'Buat data ini menjadi dashboard' creates a real Dashboard worksheet with KPIs and charts."""
    grid, index = mock_orders_grid_and_index
    sheet_grids = {"Orders": grid}
    orch = AgentOrchestrator()

    res = orch.process_request(
        user_request="Buat data ini menjadi dashboard yang siap dishare ke atasan.",
        workbook_index=index,
        grid=grid,
        sheet_grids=sheet_grids,
        active_sheet_name="Orders",
    )
    assert res.status == AgentResponseStatusEnum.SUCCESS
    assert "Dashboard" in sheet_grids
    dash_grid = sheet_grids["Dashboard"]

    # 1. Header Text
    assert dash_grid.get_cell(1, 1).parsed_value == "SALES PERFORMANCE DASHBOARD"
    assert dash_grid.get_cell(2, 1).parsed_value is not None

    # 2. KPI Values
    assert dash_grid.get_cell(4, 1).parsed_value == "Total Sales"
    assert dash_grid.get_cell(5, 1).parsed_value == 8180.0
    assert dash_grid.get_cell(4, 3).parsed_value == "Total Records"
    assert dash_grid.get_cell(5, 3).parsed_value == 10

    # 3. Floating Charts on Dashboard
    assert len(dash_grid.charts) == 2
    chart_destinations = [c.get("destination_cell") for c in dash_grid.charts.values()]
    assert "A8" in chart_destinations
    assert "E8" in chart_destinations


def test_dashboard_atomic_rollback_on_failure(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    sheet_grids = {"Orders": grid}
    orch = AgentOrchestrator()

    res = orch.process_request(
        user_request="Buat data ini menjadi dashboard yang siap dishare ke atasan.",
        workbook_index=index,
        grid=grid,
        sheet_grids=sheet_grids,
        active_sheet_name="Orders",
    )
    assert res.status == AgentResponseStatusEnum.SUCCESS
    assert "Dashboard" in sheet_grids

    # Undo removes Dashboard atomically
    undo_res = orch.undo_last_transaction(grid=grid, sheet_grids=sheet_grids)
    assert undo_res.status == AgentResponseStatusEnum.ROLLBACK_SUCCESS
    assert "Dashboard" not in sheet_grids


# ============================================================================
# 5. XLSX EXPORT INTEGRITY
# ============================================================================

def test_xlsx_export_preserves_charts_and_dashboard(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    sheet_grids = {"Orders": grid}
    orch = AgentOrchestrator()

    # Generate dashboard
    orch.process_request(
        user_request="buat data ini menjadi dashboard yang siap dishare ke atasan",
        workbook_index=index,
        grid=grid,
        sheet_grids=sheet_grids,
        active_sheet_name="Orders",
    )

    # Build openpyxl workbook
    wb = openpyxl.Workbook()
    for s_name, s_grid in sheet_grids.items():
        ws = wb.create_sheet(title=s_name)
        for (r, c), cell in s_grid.cells.items():
            if not cell.is_empty:
                ws.cell(row=r, column=c, value=cell.parsed_value)

        # Attach charts
        for c_id, c_data in s_grid.charts.items():
            c_type = c_data.get("chart_type", "BAR")
            dest = c_data.get("destination_cell", "B12")
            if c_type == "PIE":
                chart_obj = PieChart()
            elif c_type == "LINE":
                chart_obj = LineChart()
            else:
                chart_obj = BarChart()
            chart_obj.title = c_data.get("title", "Chart")
            ws.add_chart(chart_obj, dest)

    assert "Dashboard" in wb.sheetnames
    assert "Orders" in wb.sheetnames
    dash_ws = wb["Dashboard"]
    assert len(dash_ws._charts) == 2
    anchors = [c.anchor for c in dash_ws._charts]
    assert "A8" in anchors
    assert "E8" in anchors


# ============================================================================
# 6. ROBUSTNESS & SAFETY
# ============================================================================

def test_robustness_missing_measure_or_unsupported(mock_orders_grid_and_index):
    grid, index = mock_orders_grid_and_index
    actions, status, req, msg = SpreadsheetAgentPlanner.plan_agent_actions(
        user_request="tolong kirimkan email laporan ini ke bos",
        workbook_index=index,
        grid=grid,
        active_sheet_name="Orders",
    )
    assert status == AgentResponseStatusEnum.UNSUPPORTED

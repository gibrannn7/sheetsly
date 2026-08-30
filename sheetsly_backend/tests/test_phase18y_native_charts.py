"""
Comprehensive Test Suite for Phase 18.Y:
Native Worksheet Chart Rendering, Pie Chart Label Bounds, and XLSX Native Chart Data.
"""

import io
import pytest
import openpyxl
from openpyxl.chart import PieChart, BarChart, LineChart, AreaChart, Reference

from app.engine.agent.action_model import ActionTypeEnum
from app.engine.agent.agent_orchestrator import AgentOrchestrator
from app.engine.agent.agent_planner import SpreadsheetAgentPlanner
from app.engine.agent.transaction_model import AgentResponseStatusEnum
from app.engine.parser.sheet_reader import RawSheetGrid
from app.engine.profiler.workbook_index import (
    ColumnIndexEntry,
    SheetIndexEntry,
    TableIndexEntry,
    WorkbookMetadataIndex,
)
from app.engine.visualization.chart_model import ChartTypeEnum, ChartSeriesSpec
from app.engine.visualization.renderer import ChartRenderer
from app.models.schemas import CellCoordinate, CellData, DataTypeEnum, SemanticTypeEnum


@pytest.fixture
def superstore_sample_workbook():
    cells = {}
    headers = ["Order ID", "Region", "Category", "Sales", "Profit"]
    types = [DataTypeEnum.STRING, DataTypeEnum.STRING, DataTypeEnum.STRING, DataTypeEnum.FLOAT, DataTypeEnum.FLOAT]
    for c_idx, (h, dt) in enumerate(zip(headers, types), start=1):
        col_letter = chr(64 + c_idx)
        cells[(1, c_idx)] = CellData(
            coordinate=CellCoordinate(row=1, column=c_idx, cell_ref=f"{col_letter}1", col_letter=col_letter),
            original_value=h,
            parsed_value=h,
            data_type=dt,
            is_empty=False,
        )

    rows_data = [
        ("CA-101", "West", "Technology", 100.0, 20.0),
        ("CA-102", "West", "Furniture", 50.0, 5.0),
        ("CA-103", "East", "Furniture", 200.0, 30.0),
        ("CA-104", "East", "Technology", 80.0, 15.0),
        ("CA-105", "South", "Office Supplies", 30.0, 2.0),
        ("CA-106", "Central", "Office Supplies", 5.0, 1.0),
    ]

    for r_idx, row in enumerate(rows_data, start=2):
        for c_idx, val in enumerate(row, start=1):
            col_letter = chr(64 + c_idx)
            dt = types[c_idx - 1]
            cells[(r_idx, c_idx)] = CellData(
                coordinate=CellCoordinate(row=r_idx, column=c_idx, cell_ref=f"{col_letter}{r_idx}", col_letter=col_letter),
                original_value=str(val),
                parsed_value=val,
                data_type=dt,
                is_empty=False,
            )

    grid = RawSheetGrid(
        sheet_name="Orders",
        min_row=1,
        max_row=7,
        min_col=1,
        max_col=5,
        cells=cells,
    )

    columns = [
        ColumnIndexEntry(index=0, name="Order ID", normalized_name="order id", source_column_letter="A", semantic_type=SemanticTypeEnum.IDENTIFIER, data_type=DataTypeEnum.STRING),
        ColumnIndexEntry(index=1, name="Region", normalized_name="region", source_column_letter="B", semantic_type=SemanticTypeEnum.CATEGORICAL, data_type=DataTypeEnum.STRING),
        ColumnIndexEntry(index=2, name="Category", normalized_name="category", source_column_letter="C", semantic_type=SemanticTypeEnum.CATEGORICAL, data_type=DataTypeEnum.STRING),
        ColumnIndexEntry(index=3, name="Sales", normalized_name="sales", source_column_letter="D", semantic_type=SemanticTypeEnum.NUMERIC_MEASURE, data_type=DataTypeEnum.FLOAT),
        ColumnIndexEntry(index=4, name="Profit", normalized_name="profit", source_column_letter="E", semantic_type=SemanticTypeEnum.NUMERIC_MEASURE, data_type=DataTypeEnum.FLOAT),
    ]

    table_entry = TableIndexEntry(
        table_id="tbl_orders",
        sheet_name="Orders",
        name="Orders_Table",
        range_address="A1:E7",
        header_range="A1:E1",
        data_range="A2:E7",
        header_row_index=1,
        row_count=6,
        column_count=5,
        columns=columns,
    )

    sheet_entry = SheetIndexEntry(
        name="Orders",
        index=0,
        total_rows=7,
        total_columns=5,
        used_range="A1:E7",
        tables=[table_entry],
    )

    wb_index = WorkbookMetadataIndex(
        dataset_id="ds_test_p18y_superstore",
        filename="Superstore_P18Y.xlsx",
        sheet_count=1,
        sheet_names=["Orders"],
        active_sheet_name="Orders",
        sheets={"Orders": sheet_entry},
    )

    return wb_index, grid


class TestPhase18YNativeChartRendering:
    """Tests 1-8: Chart creation, anchor resolution, deterministic summary data, and undo/redo."""

    def test_create_pie_chart_anchor_b12(self, superstore_sample_workbook):
        wb_index, grid = superstore_sample_workbook
        orchestrator = AgentOrchestrator()
        res = orchestrator.process_request(
            "buatkan pie chart di B12 untuk Region berdasarkan Sales",
            wb_index,
            grid,
        )
        assert res.status == AgentResponseStatusEnum.SUCCESS
        assert "B12" in res.message
        assert len(grid.charts) == 1

        chart_spec = next(iter(grid.charts.values()))
        assert chart_spec["chart_type"] == "PIE"
        assert chart_spec["destination_cell"] == "B12"
        assert chart_spec["dimension_column"] == "Region"
        assert chart_spec["measure_column"] == "Sales"
        assert len(chart_spec["summary_data"]) > 0

    def test_pie_chart_label_and_small_slice_handling(self, tmp_path):
        out_png = tmp_path / "pie_bounds.png"
        cats = ["West", "East", "Central", "South", "TinySlice"]
        vals = [150.0, 280.0, 50.0, 30.0, 2.0]
        series = [ChartSeriesSpec(name="Sales", values=vals)]

        b64 = ChartRenderer.render_to_file(
            chart_type=ChartTypeEnum.PIE,
            title="Sales by Region with Small Slice",
            x_categories=cats,
            series=series,
            x_axis_label="Region",
            y_axis_label="Sales",
            output_file_path=out_png,
            include_base64=True,
        )
        assert out_png.exists()
        assert b64 is not None and len(b64) > 100

    def test_all_7_chart_types_headless_render(self, tmp_path):
        chart_types = [
            ChartTypeEnum.PIE,
            ChartTypeEnum.BAR,
            ChartTypeEnum.COLUMN,
            ChartTypeEnum.LINE,
            ChartTypeEnum.AREA,
            ChartTypeEnum.SCATTER,
            ChartTypeEnum.HISTOGRAM,
        ]
        cats = ["Cat A", "Cat B", "Cat C", "Cat D"]
        vals = [10.0, 20.0, 30.0, 40.0]
        series = [ChartSeriesSpec(name="Metric", values=vals)]

        for ct in chart_types:
            out_file = tmp_path / f"chart_{ct.value}.png"
            b64 = ChartRenderer.render_to_file(
                chart_type=ct,
                title=f"Test {ct.value}",
                x_categories=cats,
                series=series,
                x_axis_label="Category",
                y_axis_label="Value",
                output_file_path=out_file,
                include_base64=True,
            )
            assert out_file.exists()
            assert b64 is not None and len(b64) > 50

    def test_chart_undo_and_redo_lifecycle(self, superstore_sample_workbook):
        wb_index, grid = superstore_sample_workbook
        orchestrator = AgentOrchestrator()

        # Step 1: Create chart at B12
        orchestrator.process_request("buatkan pie chart di B12 untuk Region berdasarkan Sales", wb_index, grid)
        assert len(grid.charts) == 1

        # Step 2: Natural language undo
        res_undo = orchestrator.process_request("please undo", wb_index, grid)
        assert res_undo.status == AgentResponseStatusEnum.ROLLBACK_SUCCESS
        assert len(grid.charts) == 0

        # Step 3: Natural language redo
        res_redo = orchestrator.process_request("redo", wb_index, grid)
        assert res_redo.status == AgentResponseStatusEnum.SUCCESS
        assert len(grid.charts) == 1
        restored = next(iter(grid.charts.values()))
        assert restored["destination_cell"] == "B12"


class TestPhase18YXLSXNativeChartExport:
    """Tests 9-12: Native openpyxl chart export with attached series data."""

    def test_openpyxl_export_attaches_series_data_to_chart(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"

        # Populate sample table data
        ws["A1"] = "Order ID"
        ws["B1"] = "Region"
        ws["C1"] = "Sales"
        rows = [("101", "West", 150), ("102", "East", 280), ("103", "South", 30)]
        for r, (oid, reg, s) in enumerate(rows, start=2):
            ws[f"A{r}"] = oid
            ws[f"B{r}"] = reg
            ws[f"C{r}"] = s

        # Summary data & openpyxl chart creation
        summary_data = [{"category": "West", "value": 150}, {"category": "East", "value": 280}, {"category": "South", "value": 30}]
        aux_col = max(ws.max_column + 2, 26)
        ws.cell(row=1, column=aux_col, value="Region")
        ws.cell(row=1, column=aux_col + 1, value="Sales")
        for s_idx, item in enumerate(summary_data, start=2):
            ws.cell(row=s_idx, column=aux_col, value=item["category"])
            ws.cell(row=s_idx, column=aux_col + 1, value=item["value"])

        chart = PieChart()
        chart.title = "Sales by Region"
        chart.width = 14
        chart.height = 7.5
        data_ref = Reference(ws, min_col=aux_col + 1, min_row=1, max_row=1 + len(summary_data))
        cats_ref = Reference(ws, min_col=aux_col, min_row=2, max_row=1 + len(summary_data))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "B12")

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        # Inspect exported workbook
        loaded_wb = openpyxl.load_workbook(buf)
        loaded_ws = loaded_wb["Orders"]
        assert len(loaded_ws._charts) == 1
        loaded_chart = loaded_ws._charts[0]
        assert len(loaded_chart.series) == 1
        assert "Sales by Region" in str(loaded_chart.title)

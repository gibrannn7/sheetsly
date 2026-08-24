"""Pytest configuration and test workbook fixture generators for the 10 structural scenarios."""

import os
from pathlib import Path
from typing import Generator
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pytest
from fastapi.testclient import TestClient

from app.main import app


@pytest.fixture(scope="session")
def client() -> Generator[TestClient, None, None]:
    with TestClient(app) as test_client:
        yield test_client


@pytest.fixture
def anyio_backend():
    return "asyncio"


@pytest.fixture(scope="session")
def fixtures_dir(tmp_path_factory) -> Path:
    """Directory holding synthetic Excel workbooks for testing."""
    fdir = tmp_path_factory.mktemp("excel_fixtures")
    return fdir


@pytest.fixture(scope="session")
def vertical_table_file(fixtures_dir: Path) -> Path:
    """1. Normal vertical Excel table with clean columns (ID, Product, Date, Quantity, Revenue, Status)."""
    file_path = fixtures_dir / "1_vertical_table.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    headers = ["Transaction_ID", "Product", "Date", "Quantity", "Revenue", "Status"]
    ws.append(headers)

    data = [
        ["TXN-001", "Laptop Pro", "2026-01-15", 2, "$2,400.00", "Completed"],
        ["TXN-002", "Wireless Mouse", "2026-01-16", 10, "$250.00", "Completed"],
        ["TXN-003", "USB-C Hub", "2026-01-17", 5, "$175.00", "Pending"],
        ["TXN-004", "4K Monitor", "2026-01-18", 1, "$450.00", "Completed"],
        ["TXN-005", "Mechanical Keyboard", "2026-01-19", 3, "$360.00", "Completed"],
    ]
    for row in data:
        ws.append(row)

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture(scope="session")
def horizontal_table_file(fixtures_dir: Path) -> Path:
    """2. Horizontal data layout where rows are metrics and columns are time periods."""
    file_path = fixtures_dir / "2_horizontal_table.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MonthlyFinancials"

    # Row 1: Month headers (Jan, Feb, Mar, Apr, May, Jun)
    ws.append(["Metric", "Jan", "Feb", "Mar", "Apr", "May", "Jun"])
    ws.append(["Revenue", 10000, 12000, 15000, 14000, 18000, 20000])
    ws.append(["COGS", 4000, 4800, 6000, 5600, 7200, 8000])
    ws.append(["Operating_Expense", 3000, 3200, 3500, 3400, 3900, 4100])
    ws.append(["Net_Profit", 3000, 4000, 5500, 5000, 6900, 7900])

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture(scope="session")
def title_metadata_table_file(fixtures_dir: Path) -> Path:
    """3. Sheet with title banner, report metadata, main table, and total footer."""
    file_path = fixtures_dir / "3_title_metadata_table.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ExecutiveReport"

    # Row 1: Report Title
    ws["A1"] = "ACME CORP - QUARTERLY PERFORMANCE REPORT"
    ws.merge_cells("A1:E1")

    # Row 2: Metadata
    ws["A2"] = "Period: Q1 2026 | Prepared by: Analytics Team"

    # Row 3: Blank separator
    # Row 4: Actual Table Headers
    ws.cell(row=4, column=1, value="Region")
    ws.cell(row=4, column=2, value="Manager")
    ws.cell(row=4, column=3, value="Target_Sales")
    ws.cell(row=4, column=4, value="Actual_Sales")
    ws.cell(row=4, column=5, value="Achievement_Rate")

    # Rows 5-8: Data
    data = [
        ["Serang", "Budi", 50000, 55000, "110%"],
        ["Jakarta", "Siti", 120000, 115000, "95.8%"],
        ["Bandung", "Agus", 80000, 84000, "105%"],
        ["Surabaya", "Dewi", 90000, 92000, "102.2%"],
    ]
    for r_idx, row in enumerate(data, start=5):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Row 9: Total Footer
    ws.cell(row=9, column=1, value="Total")
    ws.cell(row=9, column=3, value="=SUM(C5:C8)")
    ws.cell(row=9, column=4, value="=SUM(D5:D8)")

    # Row 10: Note
    ws.cell(row=10, column=1, value="* Note: Figures audited as of 2026-03-31.")

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture(scope="session")
def multi_table_file(fixtures_dir: Path) -> Path:
    """4. Single sheet containing multiple separate tables divided by blank rows and columns."""
    file_path = fixtures_dir / "4_multi_table.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    # Table 1: Top Table (Rows 1-3, Cols A-C)
    ws["A1"] = "Product"
    ws["B1"] = "Category"
    ws["C1"] = "Inventory_Count"
    ws["A2"] = "Widget A"
    ws["B2"] = "Hardware"
    ws["C2"] = 150
    ws["A3"] = "Widget B"
    ws["B3"] = "Hardware"
    ws["C3"] = 80

    # Blank row 4, 5
    # Table 2: Bottom Table (Rows 6-8, Cols A-C)
    ws["A6"] = "Vendor_Name"
    ws["B6"] = "Rating"
    ws["C6"] = "Active_Orders"
    ws["A7"] = "Global Supplies"
    ws["B7"] = 4.8
    ws["C7"] = 12
    ws["A8"] = "FastTrack Logistics"
    ws["B8"] = 4.2
    ws["C8"] = 5

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture(scope="session")
def multi_row_header_file(fixtures_dir: Path) -> Path:
    """5. Table with 2-level multi-row header."""
    file_path = fixtures_dir / "5_multi_row_header.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MultiHeader"

    # Header Row 1
    ws["A1"] = "Region"
    ws["B1"] = "2025 Performance"
    ws["D1"] = "2026 Performance"
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:E1")

    # Header Row 2
    ws["A2"] = "Name"
    ws["B2"] = "Revenue"
    ws["C2"] = "Units"
    ws["D2"] = "Revenue"
    ws["E2"] = "Units"

    # Data Rows 3-5
    ws.append(["West", 100000, 500, 120000, 580])
    ws.append(["East", 80000, 420, 95000, 490])
    ws.append(["North", 60000, 310, 75000, 380])

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture(scope="session")
def empty_gaps_file(fixtures_dir: Path) -> Path:
    """6. Table containing empty rows and empty column gaps."""
    file_path = fixtures_dir / "6_empty_gaps.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gaps"

    ws.append(["ID", "Name", "Score"])
    ws.append([101, "Alice", 95])
    ws.append([])  # empty row
    ws.append([102, "Bob", 88])
    ws.append([103, "Charlie", 92])

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture(scope="session")
def mixed_values_file(fixtures_dir: Path) -> Path:
    """7. Mixed numeric and text values in numeric column (e.g. text in Quantity)."""
    file_path = fixtures_dir / "7_mixed_values.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    ws.append(["Item_Code", "Item_Name", "Quantity", "Unit_Price"])
    ws.append(["ITM-1", "Screwdriver", 50, 12.5])
    ws.append(["ITM-2", "Hammer", "Out of Stock", 25.0])  # text in numeric column
    ws.append(["ITM-3", "Drill", 15, 89.99])
    ws.append(["ITM-4", "Wrench", "Discontinued", 18.0])  # text in numeric column
    ws.append(["ITM-5", "Pliers", 30, 15.0])

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture(scope="session")
def missing_values_file(fixtures_dir: Path) -> Path:
    """8. Table with null / missing values."""
    file_path = fixtures_dir / "8_missing_values.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"

    ws.append(["Customer_ID", "Name", "Email", "Age", "City"])
    ws.append(["CUST-1", "John Doe", "john@example.com", 34, "Jakarta"])
    ws.append(["CUST-2", "Jane Smith", None, 29, None])  # missing email & city
    ws.append(["CUST-3", "Bob Wilson", "bob@example.com", None, "Bandung"])  # missing age
    ws.append(["CUST-4", "Alice Brown", None, 41, "Surabaya"])

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture(scope="session")
def duplicate_rows_file(fixtures_dir: Path) -> Path:
    """9. Table with exact duplicate rows and duplicate IDs."""
    file_path = fixtures_dir / "9_duplicate_rows.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    ws.append(["Order_ID", "Customer", "Amount"])
    ws.append(["ORD-101", "Acme Inc", 500])
    ws.append(["ORD-102", "Globex", 750])
    ws.append(["ORD-101", "Acme Inc", 500])  # Exact duplicate row & ID
    ws.append(["ORD-103", "Soylent", 320])

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture(scope="session")
def ambiguous_layout_file(fixtures_dir: Path) -> Path:
    """10. Ambiguous layout (e.g. square matrix of numbers with numeric symmetrical headers)."""
    file_path = fixtures_dir / "10_ambiguous_layout.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matrix"

    # Symmetric matrix with numeric headers and values
    ws.append([1, 2, 3, 4])
    ws.append([2, 5, 8, 9])
    ws.append([3, 8, 4, 7])
    ws.append([4, 9, 7, 6])

    wb.save(str(file_path))
    wb.close()
    return file_path

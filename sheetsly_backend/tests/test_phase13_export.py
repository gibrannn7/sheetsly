"""Unit tests for Phase 13: Full-Fidelity Excel (.xlsx) and CSV (.csv) Export."""

import pytest
import io
import openpyxl
from fastapi.testclient import TestClient

from app.main import app
from app.engine.pipeline import ingestion_pipeline
from app.engine.parser.sheet_reader import RawSheetGrid
from app.engine.agent.action_model import SpreadsheetAction, ActionTypeEnum
from app.models.schemas import CellCoordinate, CellData, DataTypeEnum

client = TestClient(app)


def test_export_xlsx_and_csv_flow(tmp_path):
    # 1. Create a 2-sheet workbook
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sales"
    ws1.append(["Item", "Price", "Qty"])
    ws1.append(["Widget A", 10.0, 5])
    ws1.append(["Widget B", 20.0, 3])
    
    ws2 = wb.create_sheet(title="Costs")
    ws2.append(["Category", "Amount"])
    ws2.append(["Rent", 1000.0])
    ws2.append(["Salaries", 2500.0])
    
    file_path = tmp_path / "multi_export.xlsx"
    wb.save(file_path)
    
    with open(file_path, "rb") as f:
        up_res = client.post("/api/v1/datasets/upload", files={"file": ("multi_export.xlsx", f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert up_res.status_code == 201
    dataset_id = up_res.json()["dataset_id"]
    
    # 2. Mutate sheet with AI Agent formula: =SUM(B2:B3) at B4
    act_res = client.post("/api/v1/agent/action", json={
        "dataset_id": dataset_id,
        "user_request": "buatkan total price",
        "active_sheet_name": "Sales"
    })
    assert act_res.status_code == 200
    assert act_res.json()["status"] == "SUCCESS"
    
    # 3. Export XLSX
    export_xlsx_res = client.get(f"/api/v1/datasets/{dataset_id}/export?format=xlsx")
    assert export_xlsx_res.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in export_xlsx_res.headers["Content-Type"]
    
    # Verify exported XLSX contents
    exported_wb = openpyxl.load_workbook(io.BytesIO(export_xlsx_res.content))
    assert set(exported_wb.sheetnames) == {"Sales", "Costs"}
    
    ws_sales = exported_wb["Sales"]
    # Check that formula is preserved as '=SUM(B2:B3)'
    found_formula = False
    for r in range(1, 10):
        for c in range(1, 10):
            val = ws_sales.cell(row=r, column=c).value
            if isinstance(val, str) and val.startswith("=SUM"):
                found_formula = True
                break
    assert found_formula is True, "Exported XLSX must preserve formula string"
    
    # 4. Export CSV (active sheet = Costs)
    export_csv_res = client.get(f"/api/v1/datasets/{dataset_id}/export?format=csv&sheet_name=Costs")
    assert export_csv_res.status_code == 200
    assert "text/csv" in export_csv_res.headers["Content-Type"]
    csv_text = export_csv_res.content.decode("utf-8")
    assert "Rent" in csv_text
    assert "Salaries" in csv_text
